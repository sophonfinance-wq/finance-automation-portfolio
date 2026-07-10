"""Source-file ingestion for the cash-completeness engine.

This module turns raw source files into the shared contracts in
:mod:`ccengine.models`:

* :func:`load_registers` reads a directory of bank-register CSVs -- one file
  per account -- and returns the reconciliation *population* as a list of
  :class:`~ccengine.models.RegisterAccount` objects. This list is the
  bank-side truth every downstream scope statement is footed against.
* :func:`load_trial_balance` reads the trial-balance CSV and returns the
  book-side claim as a list of :class:`~ccengine.models.TBRow` objects,
  preserving mis-keyed GL cells verbatim in ``gl_raw`` (typo lines are part
  of the lesson, so they must survive ingestion untouched).

Register file format
--------------------
One CSV per account. Metadata header lines start with ``#`` and use
``key: value`` (or ``key = value``) pairs::

    # entity: Juniper 42 Development LLC
    # bank: First Legacy Bank
    # bank_account_no: 000111222
    # gl_account: 615-001-00-1133
    # status: closed
    # as_of: 2026-06-30
    date,description,amount,running_balance,counterparty
    2026-06-01,Opening balance,0.00,52140.25,
    2026-06-28,"Transfer to Union National Bank",-52140.25,0.00,Union National Bank

The ``date,description,...`` column-header row is optional. ``as_of`` and
``balance`` metadata are optional; when absent they default to the date and
running balance of the last transaction. Blank lines, UTF-8 byte-order
marks, and quoted commas are all tolerated.

Trial-balance file format
-------------------------
A single CSV with columns ``sheet,gl_account,title,balance``. A header row
and ``#`` comment lines are optional and skipped.

Optional Excel support
----------------------
:func:`load_xlsx_registers` and :func:`load_xlsx_trial_balance` mirror the
CSV loaders for ``.xlsx`` workbooks. They require :mod:`openpyxl`; when it
is not installed they raise a clear :class:`RuntimeError` instead of failing
at import time -- the core engine stays stdlib-only.

All entities, banks, account numbers and figures in the examples are
fictional.
"""

from __future__ import annotations

import csv
import os
from typing import Dict, List, Optional, Sequence, Tuple

from .models import (
    ACCOUNT_STATUSES,
    STATUS_LIVE,
    RegisterAccount,
    TBRow,
    Transaction,
)
from .normalize import normalize_gl

try:  # Optional dependency: Excel ingestion only.
    import openpyxl  # type: ignore[import-untyped]

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - environment dependent
    openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False


#: Metadata keys a register file must provide (after alias resolution).
_REQUIRED_META = ("entity", "bank", "bank_account_no", "gl_account")

#: Accepted spellings for metadata keys, mapped to their canonical name.
_META_ALIASES: Dict[str, str] = {
    "entity": "entity",
    "bank": "bank",
    "bank_name": "bank",
    "bank_account_no": "bank_account_no",
    "bank_account": "bank_account_no",
    "account_no": "bank_account_no",
    "gl_account": "gl_account",
    "gl": "gl_account",
    "status": "status",
    "as_of": "as_of",
    "balance": "balance",
}


# ---------------------------------------------------------------------------
# Cell-level helpers
# ---------------------------------------------------------------------------


def _parse_amount(value: object, context: str) -> float:
    """Parse a monetary cell into a float.

    Tolerant of currency symbols, thousands separators, surrounding
    whitespace, and accounting-style parentheses for negatives. Numeric
    inputs (from Excel cells) pass straight through.

    Parameters
    ----------
    value:
        The raw cell value.
    context:
        Where the cell came from (file / row), used in error messages.

    Raises
    ------
    ValueError
        If the cell is blank or not a number -- amounts are never guessed.
    """
    if isinstance(value, bool):
        raise ValueError(f"{context}: expected an amount, got boolean {value!r}")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{context}: amount cell is blank")
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    text = text.replace("$", "").replace(",", "").strip()
    try:
        amount = float(text)
    except ValueError:
        raise ValueError(f"{context}: cannot parse amount {value!r}") from None
    return -amount if negative else amount


def _parse_meta_line(text: str) -> Optional[Tuple[str, str]]:
    """Parse one ``# key: value`` metadata line into ``(canonical_key, value)``.

    Returns ``None`` for comment lines that are not key/value pairs, and for
    keys the loader does not recognise (unknown keys are ignored rather than
    fatal so register exports can carry extra annotations).
    """
    body = text.lstrip("#").strip()
    sep = ":" if ":" in body else ("=" if "=" in body else None)
    if sep is None:
        return None
    key, _, value = body.partition(sep)
    canonical = _META_ALIASES.get(key.strip().lower().replace(" ", "_"))
    if canonical is None:
        return None
    return canonical, value.strip()


def _row_is_blank(row: Sequence[object]) -> bool:
    """True when a CSV/xlsx row carries no content at all."""
    return not row or not any(str(cell or "").strip() for cell in row)


def _cell_str(value: object) -> str:
    """Render an Excel/CSV cell as a clean string (dates -> ISO)."""
    if value is None:
        return ""
    iso = getattr(value, "isoformat", None)
    if callable(iso):  # datetime.date / datetime.datetime cells
        return iso()[:10]
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# ---------------------------------------------------------------------------
# Row assembly (shared by CSV and xlsx paths)
# ---------------------------------------------------------------------------


def _build_transaction(row: Sequence[object], context: str) -> Transaction:
    """Turn one data row into a :class:`Transaction`.

    Expects ``date, description, amount, running_balance[, counterparty]``.
    """
    if len(row) < 4:
        raise ValueError(
            f"{context}: transaction row needs at least 4 columns "
            f"(date, description, amount, running_balance), got {list(row)!r}"
        )
    counterparty = _cell_str(row[4]) if len(row) > 4 else ""
    return Transaction(
        date=_cell_str(row[0]),
        description=_cell_str(row[1]),
        amount=_parse_amount(row[2], f"{context} amount"),
        running_balance=_parse_amount(row[3], f"{context} running_balance"),
        counterparty=counterparty or None,
    )


def _build_register_account(
    source_file: str,
    meta: Dict[str, str],
    transactions: List[Transaction],
) -> RegisterAccount:
    """Assemble a :class:`RegisterAccount` from parsed metadata and rows."""
    missing = [key for key in _REQUIRED_META if not meta.get(key, "").strip()]
    if missing:
        raise ValueError(
            f"{source_file}: register file is missing required metadata "
            f"{missing}; expected '# key: value' header lines"
        )

    status = (meta.get("status") or STATUS_LIVE).strip().lower()
    if status not in ACCOUNT_STATUSES:
        raise ValueError(
            f"{source_file}: status must be one of {ACCOUNT_STATUSES}, "
            f"got {meta.get('status')!r}"
        )

    if "balance" in meta:
        balance = _parse_amount(meta["balance"], f"{source_file} balance metadata")
    elif transactions:
        balance = transactions[-1].running_balance
    else:
        balance = 0.0

    as_of = meta.get("as_of", "").strip()
    if not as_of and transactions:
        as_of = transactions[-1].date

    gl_raw = meta["gl_account"]
    return RegisterAccount(
        source_file=source_file,
        entity=meta["entity"],
        bank=meta["bank"],
        bank_account_no=meta["bank_account_no"],
        gl_raw=gl_raw,
        gl_norm=normalize_gl(gl_raw),
        balance=round(balance, 2),
        as_of=as_of,
        status=status,
        transactions=transactions,
    )


def _build_tb_row(source_file: str, row: Sequence[object], context: str) -> TBRow:
    """Turn one TB data row (``sheet,gl_account,title,balance``) into a :class:`TBRow`."""
    if len(row) < 4:
        raise ValueError(
            f"{context}: trial-balance row needs 4 columns "
            f"(sheet, gl_account, title, balance), got {list(row)!r}"
        )
    gl_raw = _cell_str(row[1])
    return TBRow(
        source_file=source_file,
        sheet=_cell_str(row[0]),
        gl_raw=gl_raw,
        gl_norm=normalize_gl(gl_raw) if gl_raw else "",
        title=_cell_str(row[2]),
        balance=_parse_amount(row[3], f"{context} balance"),
    )


# ---------------------------------------------------------------------------
# CSV loaders (core, stdlib-only)
# ---------------------------------------------------------------------------


def load_registers(registers_dir: str) -> List[RegisterAccount]:
    """Load every register CSV in a directory into the population of record.

    Files are read in sorted name order so the population is deterministic.
    Blank lines are skipped, a UTF-8 BOM is tolerated, quoted commas inside
    descriptions/counterparties are handled by the ``csv`` module, and the
    optional ``date,description,...`` header row is recognised and skipped.

    Parameters
    ----------
    registers_dir:
        Directory containing one ``*.csv`` register file per bank account.

    Returns
    -------
    list[RegisterAccount]
        One account per file, in sorted file order.

    Raises
    ------
    FileNotFoundError
        If ``registers_dir`` does not exist or is not a directory.
    ValueError
        On malformed files: missing required metadata, unparseable amounts,
        short transaction rows, or an invalid status. Bad data is a loud
        failure, never a silent skip -- a skipped register file would
        understate the population and defeat the completeness check.
    """
    if not os.path.isdir(registers_dir):
        raise FileNotFoundError(f"register directory not found: {registers_dir}")

    accounts: List[RegisterAccount] = []
    names = sorted(n for n in os.listdir(registers_dir) if n.lower().endswith(".csv"))
    for name in names:
        path = os.path.join(registers_dir, name)
        meta: Dict[str, str] = {}
        transactions: List[Transaction] = []
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                if _row_is_blank(row):
                    continue
                first = str(row[0]).strip()
                if first.startswith("#"):
                    parsed = _parse_meta_line(",".join(str(c) for c in row))
                    if parsed:
                        meta[parsed[0]] = parsed[1]
                    continue
                if first.lower() == "date":  # optional column-header row
                    continue
                context = f"{name} line {len(transactions) + 1}"
                transactions.append(_build_transaction(row, context))
        accounts.append(_build_register_account(name, meta, transactions))
    return accounts


def load_trial_balance(tb_path: str) -> List[TBRow]:
    """Load the trial-balance CSV (``sheet,gl_account,title,balance``).

    The optional header row and any ``#`` comment lines are skipped; blank
    lines and a UTF-8 BOM are tolerated. Mis-keyed GL cells are *kept*: they
    normalise to whatever they are (or ``""`` when blank) and flow downstream
    where the classifier and verifier can flag them -- ingestion never
    launders a typo line out of the data.

    Parameters
    ----------
    tb_path:
        Path to the trial-balance CSV file.

    Returns
    -------
    list[TBRow]
        One row per TB data line, in file order.

    Raises
    ------
    FileNotFoundError
        If ``tb_path`` does not exist.
    ValueError
        On rows with fewer than 4 columns or unparseable balances.
    """
    if not os.path.isfile(tb_path):
        raise FileNotFoundError(f"trial balance not found: {tb_path}")

    source_file = os.path.basename(tb_path)
    rows: List[TBRow] = []
    with open(tb_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if _row_is_blank(row):
                continue
            first = str(row[0]).strip()
            if first.startswith("#"):
                continue
            if first.lower() == "sheet":  # optional header row
                continue
            context = f"{source_file} line {reader.line_num}"
            rows.append(_build_tb_row(source_file, row, context))
    return rows


# ---------------------------------------------------------------------------
# Optional xlsx loaders (require openpyxl)
# ---------------------------------------------------------------------------


def _require_openpyxl(caller: str) -> None:
    """Raise a clear error when an xlsx loader is called without openpyxl."""
    if not _HAS_OPENPYXL:
        raise RuntimeError(
            f"{caller} requires the optional dependency 'openpyxl' "
            "(pip install openpyxl); the CSV loaders need only the stdlib"
        )


def load_xlsx_registers(registers_dir: str) -> List[RegisterAccount]:
    """Excel variant of :func:`load_registers` (requires ``openpyxl``).

    Reads every ``*.xlsx`` workbook in ``registers_dir`` (first worksheet
    of each), using the same layout as the CSV format: ``# key: value``
    metadata lines in column A -- or plain ``key`` / ``value`` pairs in
    columns A and B -- followed by transaction rows. Date cells stored as
    real Excel dates are converted to ISO strings.

    Raises
    ------
    RuntimeError
        If ``openpyxl`` is not installed.
    """
    _require_openpyxl("load_xlsx_registers")
    if not os.path.isdir(registers_dir):
        raise FileNotFoundError(f"register directory not found: {registers_dir}")

    accounts: List[RegisterAccount] = []
    names = sorted(n for n in os.listdir(registers_dir) if n.lower().endswith(".xlsx"))
    for name in names:
        path = os.path.join(registers_dir, name)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            meta: Dict[str, str] = {}
            transactions: List[Transaction] = []
            for cells in sheet.iter_rows(values_only=True):
                if _row_is_blank(cells):
                    continue
                first = _cell_str(cells[0])
                if first.startswith("#"):
                    line = first
                    if ":" not in line and "=" not in line and len(cells) > 1:
                        line = f"{first}: {_cell_str(cells[1])}"
                    parsed = _parse_meta_line(line)
                    if parsed:
                        meta[parsed[0]] = parsed[1]
                    continue
                key = _META_ALIASES.get(first.lower().replace(" ", "_"))
                if key and not transactions and len(cells) > 1:
                    meta[key] = _cell_str(cells[1])
                    continue
                if first.lower() == "date":
                    continue
                context = f"{name} row {len(transactions) + 1}"
                transactions.append(_build_transaction(list(cells), context))
        finally:
            workbook.close()
        accounts.append(_build_register_account(name, meta, transactions))
    return accounts


def load_xlsx_trial_balance(tb_path: str) -> List[TBRow]:
    """Excel variant of :func:`load_trial_balance` (requires ``openpyxl``).

    Every worksheet is read. Two layouts are accepted per sheet:

    * four columns ``sheet, gl_account, title, balance`` (same as the CSV);
    * three columns ``gl_account, title, balance``, in which case the
      worksheet's own title is used as the ``sheet`` value.

    A header row (first cell ``sheet`` or ``gl_account``, case-insensitive)
    is recognised and skipped.

    Raises
    ------
    RuntimeError
        If ``openpyxl`` is not installed.
    """
    _require_openpyxl("load_xlsx_trial_balance")
    if not os.path.isfile(tb_path):
        raise FileNotFoundError(f"trial balance not found: {tb_path}")

    source_file = os.path.basename(tb_path)
    rows: List[TBRow] = []
    workbook = openpyxl.load_workbook(tb_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for index, cells in enumerate(sheet.iter_rows(values_only=True), start=1):
                if _row_is_blank(cells):
                    continue
                first = _cell_str(cells[0]).lower()
                if first.startswith("#") or first in ("sheet", "gl_account"):
                    continue
                row = list(cells)
                if len(row) < 4 or (len(row) >= 3 and first != "" and row[3] is None):
                    # Three-column layout: prepend the worksheet title.
                    row = [sheet.title] + row
                context = f"{source_file}[{sheet.title}] row {index}"
                rows.append(_build_tb_row(source_file, row, context))
    finally:
        workbook.close()
    return rows

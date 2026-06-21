"""Integrity-preserving edits for openpyxl-built workbooks.

openpyxl silently drops parts it does not model when it saves. The two that
matter most for a review-grade workpaper are **threaded comments**
(``xl/threadedComments/*`` plus ``xl/persons/*``) and **embedded images**
(``xl/media/*``). A naive ``load_workbook`` / ``save`` round trip to change one
label therefore destroys reviewer comments and evidence screenshots with no
warning and no error.

This module edits at the OOXML (zip/XML) layer instead: every part it does not
deliberately touch is copied byte-for-byte, so custom parts survive. Around the
edit it provides the guards a shippable workpaper needs:

* ``preflight``   - lock-file detection and zip-integrity test
* ``make_backup`` - timestamped, validated backup before any write
* anchor locking  - named output cells must hold their value across the edit
* ``validate_unchanged`` - part-inventory diff, XML well-formedness, reload,
  and anchor comparison, with automatic rollback on any failure

Stdlib + openpyxl only. Generic by design - it carries no engagement data.
"""

from __future__ import annotations

import datetime
import glob
import os
import re
import shutil
import zipfile
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Set, Tuple
from xml.dom import minidom
from xml.sax.saxutils import escape

import openpyxl

Anchor = Tuple[str, str]  # (sheet_name, cell_ref)

_SHARED_STRINGS = "xl/sharedStrings.xml"
_WORKSHEET_RE = re.compile(r"xl/worksheets/sheet\d+\.xml$")


class SafeEditError(Exception):
    """Raised when an integrity-preserving edit cannot complete safely."""


@dataclass
class PreflightResult:
    ok: bool
    lock_files: List[str]
    zip_ok: bool


@dataclass
class EditReport:
    backup: str
    replacements: int
    parts_preserved: int
    anchors_checked: int
    issues: List[str] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Pre-flight
# --------------------------------------------------------------------------- #
def find_lock_files(path: str) -> List[str]:
    """Return any Excel lock files (``~$*.xlsx``) sitting beside ``path``."""
    directory = os.path.dirname(os.path.abspath(path)) or "."
    return sorted(glob.glob(os.path.join(directory, "~$*.xlsx")))


def zip_ok(path: str) -> bool:
    """True when the file is a structurally valid zip (``testzip`` is None)."""
    try:
        with zipfile.ZipFile(path) as z:
            return z.testzip() is None
    except zipfile.BadZipFile:
        return False


def preflight(path: str) -> PreflightResult:
    """Check for open-file locks and zip corruption before touching a file."""
    locks = find_lock_files(path)
    ok_zip = zip_ok(path)
    return PreflightResult(ok=(not locks and ok_zip), lock_files=locks, zip_ok=ok_zip)


def make_backup(path: str, tag: str = "edit", timestamp: Optional[str] = None) -> str:
    """Copy ``path`` to a timestamped backup and verify it is a valid zip."""
    stamp = timestamp or datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = "%s.backup_before_%s_%s.xlsx" % (path, tag, stamp)
    shutil.copy2(path, backup)
    if not zip_ok(backup):
        raise SafeEditError("backup is not a valid zip: %s" % backup)
    return backup


# --------------------------------------------------------------------------- #
# Inspection
# --------------------------------------------------------------------------- #
def list_parts(path: str) -> Set[str]:
    """Return the set of archive member names (the OOXML part inventory)."""
    with zipfile.ZipFile(path) as z:
        return set(z.namelist())


def read_anchors(path: str, anchors: Sequence[Anchor]) -> Dict[Anchor, object]:
    """Read the cached value of each ``(sheet, cell)`` anchor."""
    if not anchors:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        return {(s, c): wb[s][c].value for s, c in anchors}
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# The surgical edit
# --------------------------------------------------------------------------- #
def _rewrite_zip(src: str, replacements: Dict[str, bytes], dst: str) -> None:
    """Copy every part of ``src`` to ``dst`` byte-for-byte, overriding the
    members named in ``replacements`` with new bytes. Part order is preserved.
    """
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
        dst, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = replacements.get(item.filename, zin.read(item.filename))
            info = zipfile.ZipInfo(item.filename, date_time=item.date_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            info.internal_attr = item.internal_attr
            info.create_system = item.create_system
            zout.writestr(info, data)


def replace_label_text(path: str, old_text: str, new_text: str) -> int:
    """Replace a cell label's text in place, preserving every other part.

    Edits the ``<t>...</t>`` runs that carry the text wherever Excel stored it -
    the shared-strings table (``xl/sharedStrings.xml``, the usual Excel layout)
    and/or inline strings inside the worksheet parts (the layout openpyxl
    writes). Comment, person, and media parts are never scanned, so reviewer
    comments and images are untouched. Returns the number of runs replaced;
    raises if the text is not present.

    Limitation: matches single-run ``<t>...</t>`` cells (the usual case for a
    label). Rich-text strings split across multiple runs are not handled.
    """
    pattern = re.compile(r"<t( [^>]*)?>" + re.escape(escape(old_text)) + r"</t>")
    esc_new = escape(new_text)

    changed: Dict[str, bytes] = {}
    total = 0
    with zipfile.ZipFile(path) as zin:
        targets = [
            n for n in zin.namelist()
            if n == _SHARED_STRINGS or _WORKSHEET_RE.match(n)
        ]
        for name in targets:
            text = zin.read(name).decode("utf-8")
            hits = len(pattern.findall(text))
            if hits:
                total += hits
                changed[name] = pattern.sub(
                    lambda m: "<t%s>%s</t>" % (m.group(1) or "", esc_new), text
                ).encode("utf-8")

    if total == 0:
        raise SafeEditError("text not found: %r" % old_text)

    tmp = path + ".tmp_safeedit"
    _rewrite_zip(path, changed, tmp)
    os.replace(tmp, path)
    return total


# --------------------------------------------------------------------------- #
# Post-edit validation
# --------------------------------------------------------------------------- #
def validate_unchanged(
    path: str,
    before_parts: Set[str],
    before_anchors: Dict[Anchor, object],
    anchors: Sequence[Anchor],
) -> List[str]:
    """Return a list of integrity issues; empty means the edit is clean."""
    issues: List[str] = []
    with zipfile.ZipFile(path) as z:
        after_parts = set(z.namelist())
        added = after_parts - before_parts
        removed = before_parts - after_parts
        if added or removed:
            issues.append("part inventory changed: added=%s removed=%s" % (sorted(added), sorted(removed)))
        if z.testzip() is not None:
            issues.append("zip integrity failed")
        for name in z.namelist():
            if name.endswith((".xml", ".rels", ".vml")):
                try:
                    minidom.parseString(z.read(name))
                except Exception as exc:  # noqa: BLE001 - report, do not raise
                    issues.append("malformed XML in %s: %s" % (name, exc))

    try:
        openpyxl.load_workbook(path).close()
    except Exception as exc:  # noqa: BLE001 - report, do not raise
        issues.append("openpyxl reload failed: %s" % exc)

    after_anchors = read_anchors(path, anchors)
    for key, value in before_anchors.items():
        if after_anchors.get(key) != value:
            issues.append("anchor %s moved: %r -> %r" % (key, value, after_anchors.get(key)))
    return issues


def safe_edit_label(
    path: str,
    old_text: str,
    new_text: str,
    anchors: Optional[Sequence[Anchor]] = None,
    tag: str = "label",
) -> EditReport:
    """Replace a shared-string label, preserving custom parts and locked anchors.

    Flow: pre-flight -> backup -> capture parts + anchors -> surgical replace ->
    validate (parts, XML, reload, anchors). On any failure the original is
    restored from the backup and ``SafeEditError`` is raised.
    """
    pre = preflight(path)
    if not pre.ok:
        raise SafeEditError(
            "preflight failed (lock_files=%s, zip_ok=%s)" % (pre.lock_files, pre.zip_ok)
        )

    anchors = list(anchors or [])
    before_parts = list_parts(path)
    before_anchors = read_anchors(path, anchors)
    backup = make_backup(path, tag=tag)

    try:
        count = replace_label_text(path, old_text, new_text)
        issues = validate_unchanged(path, before_parts, before_anchors, anchors)
        if issues:
            raise SafeEditError("; ".join(issues))
    except Exception:
        shutil.copy2(backup, path)  # roll back to the validated backup
        raise

    return EditReport(
        backup=backup,
        replacements=count,
        parts_preserved=len(before_parts),
        anchors_checked=len(anchors),
    )

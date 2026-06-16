"""Optional Excel workbook export (openpyxl).

Mirrors the Markdown lineage in a spreadsheet a preparer could open: an Evidence
sheet, a Surplus-Details sheet, and a Summary sheet, plus an FX sheet. This is a
convenience artifact; the engine and tests do not depend on it.
"""

from __future__ import annotations

from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font

from .engine import EntityYearResult
from .model import Structure

_BOLD = Font(bold=True)


def _write_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _BOLD


def build_workbook(
    results: List[EntityYearResult], structure: Structure
) -> Workbook:
    """Build an :class:`openpyxl.Workbook` from engine results."""
    wb = Workbook()

    # FX sheet (first / default).
    fx_ws = wb.active
    fx_ws.title = "FX"
    _write_header(fx_ws, ["Year", "USD->CAD"])
    for year in sorted(structure.fx.usd_cad):
        fx_ws.append([year, structure.fx.usd_cad[year]])

    # Evidence sheet.
    ev = wb.create_sheet("Evidence")
    _write_header(
        ev,
        [
            "Entity", "FY", "Standalone income", "Reg5907(2) adj", "Exempt %",
            "Distribution", "Capital contribution", "Return of capital",
        ],
    )
    for r in sorted(results, key=lambda r: (r.entity, r.year)):
        ev.append(
            [
                structure.entities[r.entity].name, r.year,
                r.standalone_taxable_income, r.reg_5907_2_adjustment,
                r.exempt_portion, r.distribution, r.capital_contribution,
                r.return_of_capital,
            ]
        )

    # Surplus-Details sheet.
    sd = wb.create_sheet("Surplus-Details")
    _write_header(
        sd,
        [
            "Entity", "FY", "Standalone surplus", "Allocable surplus",
            "Cur exempt add", "Cur taxable add", "Elevated exempt",
            "Elevated taxable", "Exempt cap", "Cap binding",
        ],
    )
    for r in sorted(results, key=lambda r: (r.entity, r.year)):
        sd.append(
            [
                structure.entities[r.entity].name, r.year, r.standalone_surplus,
                r.allocable_surplus, r.current_exempt_addition,
                r.current_taxable_addition, r.elevated_exempt, r.elevated_taxable,
                r.exempt_cap_amount, "Y" if r.exempt_cap_binding else "",
            ]
        )

    # Summary sheet (closing balances + CAD).
    sm = wb.create_sheet("Summary")
    _write_header(
        sm,
        [
            "Entity", "FY", "Cur", "Exempt", "Taxable", "Pre-acq", "ACB",
            "FX->CAD", "Total surplus (CAD)",
        ],
    )
    for r in sorted(results, key=lambda r: (r.entity, r.year)):
        total = (
            r.closing.exempt_surplus
            + r.closing.taxable_surplus
            + r.closing.pre_acquisition_capital
        )
        sm.append(
            [
                structure.entities[r.entity].name, r.year, r.currency,
                r.closing.exempt_surplus, r.closing.taxable_surplus,
                r.closing.pre_acquisition_capital, r.closing.acb,
                r.fx_rate_to_cad, round(total * r.fx_rate_to_cad, 2),
            ]
        )

    return wb

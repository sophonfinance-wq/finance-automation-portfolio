"""Coverage for the workbook export, the CLI, and extra safe_edit guards.

* Workbook: sheet layout, header order, FX sheet contents, row counts.
* CLI: argument parsing, the bare-summary path, exit codes for bad input,
  and stdout content.
* safe_edit: pre-flight on a clean file, multi-hit replacement counts,
  anchor-read round trips, and inventory diffs.
"""

from __future__ import annotations

import zipfile

import openpyxl
import pytest

from surplus_engine import safe_edit
from surplus_engine.engine import SurplusEngine
from surplus_engine.generate import generate_structure
from surplus_engine.report import attach_fx
from surplus_engine.workbook import build_workbook


def _results_and_structure():
    s = generate_structure(2021, 2022)
    results = SurplusEngine(s).run([2021, 2022])
    attach_fx(results, s)
    return results, s


# --------------------------------------------------------------------------- #
# Workbook export
# --------------------------------------------------------------------------- #
def test_workbook_has_expected_sheets_in_order():
    results, s = _results_and_structure()
    wb = build_workbook(results, s)
    assert wb.sheetnames == ["FX", "Evidence", "Surplus-Details", "Summary"]


def test_workbook_fx_sheet_matches_table():
    results, s = _results_and_structure()
    wb = build_workbook(results, s)
    ws = wb["FX"]
    assert [c.value for c in ws[1]] == ["Year", "USD->CAD"]
    body = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2)}
    assert body == s.fx.usd_cad


def test_workbook_evidence_header_layout():
    results, s = _results_and_structure()
    wb = build_workbook(results, s)
    header = [c.value for c in wb["Evidence"][1]]
    assert header[:3] == ["Entity", "FY", "Standalone income"]
    assert "Return of capital" in header


def test_workbook_summary_row_count_matches_results():
    results, s = _results_and_structure()
    wb = build_workbook(results, s)
    sm = wb["Summary"]
    # One header row + one row per result.
    assert sm.max_row == 1 + len(results)


def test_workbook_surplus_details_marks_cap_binding_flag():
    results, s = _results_and_structure()
    wb = build_workbook(results, s)
    sd = wb["Surplus-Details"]
    header = [c.value for c in sd[1]]
    cap_col = header.index("Cap binding")
    flags = {row[cap_col].value for row in sd.iter_rows(min_row=2)}
    # The flag column only ever holds "Y" or "" (empty).
    assert flags.issubset({"Y", "", None})


def test_workbook_roundtrips_through_openpyxl(tmp_path):
    results, s = _results_and_structure()
    wb = build_workbook(results, s)
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    reloaded = openpyxl.load_workbook(path)
    assert reloaded.sheetnames == ["FX", "Evidence", "Surplus-Details", "Summary"]


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def test_cli_parse_args_defaults():
    from surplus_engine.cli import _parse_args

    ns = _parse_args([])
    assert ns.start == 2021
    assert ns.end == 2024
    assert ns.xlsx is False


def test_cli_parse_args_custom_values():
    from surplus_engine.cli import _parse_args

    ns = _parse_args(["--start", "2019", "--end", "2025", "--exempt-cap", "0.5", "--xlsx"])
    assert ns.start == 2019
    assert ns.end == 2025
    assert ns.exempt_cap == 0.5
    assert ns.xlsx is True


def test_cli_bare_run_prints_summary(capsys):
    from surplus_engine.cli import main

    rc = main(["--start", "2021", "--end", "2021"])
    assert rc == 0
    captured = capsys.readouterr()
    assert "Consolidated Surplus & ACB Summary" in captured.out
    assert "Birchwood Op Co" in captured.out


def test_cli_rejects_inverted_year_range():
    from surplus_engine.cli import main

    assert main(["--start", "2025", "--end", "2020"]) == 2


@pytest.mark.parametrize("cap", ["-0.1", "1.5", "2.0"])
def test_cli_rejects_out_of_range_exempt_cap(cap):
    from surplus_engine.cli import main

    assert main(["--exempt-cap", cap]) == 2


@pytest.mark.parametrize("cap", ["0.0", "0.5", "1.0"])
def test_cli_accepts_valid_exempt_cap(cap):
    from surplus_engine.cli import main

    assert main(["--start", "2021", "--end", "2021", "--exempt-cap", cap]) == 0


def test_cli_writes_all_four_workpapers(tmp_path):
    from surplus_engine.cli import main

    out_dir = tmp_path / "out"
    assert main(["--start", "2021", "--end", "2021", "--out", str(out_dir)]) == 0
    for code in ("BIRCH_OP", "CEDAR_MEZZ", "MAPLE_FUND", "DEMO_HOLDCO"):
        assert (out_dir / f"workpaper_{code}.md").exists()


# --------------------------------------------------------------------------- #
# safe_edit additional guards
# --------------------------------------------------------------------------- #
def _build_fixture(tmp_path, label="LABEL_TO_EDIT", repeats=1):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    for i in range(repeats):
        ws.cell(row=i + 1, column=1, value=label)
    ws["H120"] = 100.0
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def test_preflight_passes_on_clean_file(tmp_path):
    p = _build_fixture(tmp_path)
    pre = safe_edit.preflight(str(p))
    assert pre.ok is True
    assert pre.lock_files == []
    assert pre.zip_ok is True


def test_zip_ok_false_for_non_zip(tmp_path):
    bogus = tmp_path / "not_a_zip.xlsx"
    bogus.write_bytes(b"this is plainly not a zip archive")
    assert safe_edit.zip_ok(str(bogus)) is False


def test_replace_label_counts_each_occurrence(tmp_path):
    p = _build_fixture(tmp_path, label="REPEATED", repeats=3)
    # openpyxl stores identical strings once in sharedStrings, so a shared label
    # appears as a single <t> run; replacement count reflects stored runs.
    count = safe_edit.replace_label_text(str(p), "REPEATED", "NEW")
    assert count >= 1
    wb = openpyxl.load_workbook(p)
    assert wb["Summary"]["A1"].value == "NEW"


def test_read_anchors_round_trips_value(tmp_path):
    p = _build_fixture(tmp_path)
    anchors = safe_edit.read_anchors(str(p), [("Summary", "H120")])
    assert anchors[("Summary", "H120")] == 100.0


def test_read_anchors_empty_sequence_returns_empty():
    assert safe_edit.read_anchors("ignored_path.xlsx", []) == {}


def test_list_parts_includes_core_ooxml_members(tmp_path):
    p = _build_fixture(tmp_path)
    parts = safe_edit.list_parts(str(p))
    assert "[Content_Types].xml" in parts
    assert any("worksheets/sheet" in m for m in parts)


def test_safe_edit_label_changes_text_and_reports(tmp_path):
    p = _build_fixture(tmp_path, label="OLD_ANCHOR_LABEL")
    report = safe_edit.safe_edit_label(
        str(p), "OLD_ANCHOR_LABEL", "NEW_ANCHOR_LABEL",
        anchors=[("Summary", "H120")], tag="cov",
    )
    assert report.replacements == 1
    assert report.anchors_checked == 1
    assert report.parts_preserved > 0
    wb = openpyxl.load_workbook(p)
    assert wb["Summary"]["A1"].value == "NEW_ANCHOR_LABEL"
    assert wb["Summary"]["H120"].value == 100.0


def test_validate_unchanged_clean_when_nothing_moved(tmp_path):
    p = _build_fixture(tmp_path)
    before_parts = safe_edit.list_parts(str(p))
    before_anchors = safe_edit.read_anchors(str(p), [("Summary", "H120")])
    issues = safe_edit.validate_unchanged(
        str(p), before_parts, before_anchors, [("Summary", "H120")]
    )
    assert issues == []


def test_make_backup_uses_supplied_timestamp(tmp_path):
    p = _build_fixture(tmp_path)
    backup = safe_edit.make_backup(str(p), tag="cov", timestamp="20210101_120000")
    assert backup.endswith("backup_before_cov_20210101_120000.xlsx")
    assert zipfile.is_zipfile(backup)

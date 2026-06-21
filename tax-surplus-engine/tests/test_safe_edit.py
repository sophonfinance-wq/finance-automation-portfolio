"""Integrity-preserving edit tests for ``surplus_engine.safe_edit``.

Proves the hazard the module exists to defeat - a naive openpyxl round trip
drops threaded comments and embedded images - and that ``safe_edit`` changes the
target label while preserving every untouched part and leaving anchor cells
unmoved. Rolls back when an anchor would move or a part would vanish. All data is
fictional; the fixtures are built in a temp dir.
"""

from __future__ import annotations

import zipfile

import openpyxl
import pytest

from surplus_engine import safe_edit

# Custom OOXML parts openpyxl does not model and silently drops on save.
CUSTOM_PARTS = {
    "xl/threadedComments/threadedComment1.xml": (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<ThreadedComments xmlns="http://schemas.microsoft.com/office/'
        b'spreadsheetml/2018/threadedcomments"/>'
    ),
    "xl/persons/person.xml": (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<personList xmlns="http://schemas.microsoft.com/office/'
        b'spreadsheetml/2018/threadedcomments"/>'
    ),
    "xl/media/image1.png": b"\x89PNG\r\n\x1a\n_fake_image_bytes_for_test_",
}

OLD_LABEL = "SOURCE_LABEL_OLD_PATH"
NEW_LABEL = "Source: fictional workpapers.xlsx, Member Capital Accounts tab"
ANCHOR = ("Summary", "H120")
ANCHOR_VALUE = 312491.58


def _build_workbook(path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = OLD_LABEL          # a shared string we will edit
    ws["H120"] = ANCHOR_VALUE     # a locked anchor (literal, not a formula)
    wb.save(path)


def _inject_custom_parts(path) -> None:
    with zipfile.ZipFile(path) as zin:
        items = [(i, zin.read(i.filename)) for i in zin.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in items:
            zout.writestr(info, data)
        for name, data in CUSTOM_PARTS.items():
            zout.writestr(name, data)


def _make_fixture(tmp_path):
    p = tmp_path / "book.xlsx"
    _build_workbook(p)
    _inject_custom_parts(p)
    return p


def _parts(path) -> set:
    with zipfile.ZipFile(path) as z:
        return set(z.namelist())


def test_openpyxl_roundtrip_drops_custom_parts(tmp_path):
    """The hazard: load + save through openpyxl destroys comments and images."""
    p = _make_fixture(tmp_path)
    assert set(CUSTOM_PARTS).issubset(_parts(p))

    wb = openpyxl.load_workbook(p)
    wb.save(p)  # naive round trip

    remaining = _parts(p)
    for name in CUSTOM_PARTS:
        assert name not in remaining


def test_safe_edit_preserves_parts_and_changes_label(tmp_path):
    p = _make_fixture(tmp_path)
    before = _parts(p)

    report = safe_edit.safe_edit_label(
        str(p), OLD_LABEL, NEW_LABEL, anchors=[ANCHOR], tag="test"
    )

    assert report.replacements == 1
    assert report.anchors_checked == 1
    # Every untouched part - including the custom ones - survived unchanged.
    assert _parts(p) == before
    for name in CUSTOM_PARTS:
        assert name in _parts(p)

    wb = openpyxl.load_workbook(p)
    assert wb["Summary"]["A1"].value == NEW_LABEL
    assert wb["Summary"]["H120"].value == ANCHOR_VALUE


def test_preflight_flags_lock_file(tmp_path):
    p = _make_fixture(tmp_path)
    (tmp_path / "~$book.xlsx").write_bytes(b"lock")
    pre = safe_edit.preflight(str(p))
    assert pre.ok is False
    assert any("~$" in f for f in pre.lock_files)


def test_make_backup_is_valid_zip(tmp_path):
    p = _make_fixture(tmp_path)
    backup = safe_edit.make_backup(str(p), tag="t", timestamp="20200101_000000")
    assert backup.endswith("20200101_000000.xlsx")
    assert safe_edit.zip_ok(backup)


def test_missing_label_raises_and_leaves_file_intact(tmp_path):
    p = _make_fixture(tmp_path)
    before_bytes = p.read_bytes()
    with pytest.raises(safe_edit.SafeEditError):
        safe_edit.safe_edit_label(str(p), "NOT_PRESENT_ANYWHERE", "x")
    assert p.read_bytes() == before_bytes


def test_validate_detects_anchor_move_and_part_loss(tmp_path):
    p = _make_fixture(tmp_path)
    before_parts = safe_edit.list_parts(str(p))
    before_anchors = safe_edit.read_anchors(str(p), [ANCHOR])

    # Mutate an anchor via a naive save (which also drops the custom parts).
    wb = openpyxl.load_workbook(p)
    wb["Summary"]["H120"] = 999999
    wb.save(p)

    issues = safe_edit.validate_unchanged(
        str(p), before_parts, before_anchors, [ANCHOR]
    )
    assert any("anchor" in i for i in issues)
    assert any("part inventory" in i for i in issues)

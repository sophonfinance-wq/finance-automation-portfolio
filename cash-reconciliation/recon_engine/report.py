"""Evidence-log rendering: 5-section Markdown and an .xlsx workbook.

The same :class:`~recon_engine.engine.ReconResult` is rendered two ways:

* :func:`render_markdown` — a committed, human-readable evidence log with the
  five canonical sections (Summary / Cash / Debt / Flagged / Notes).
* :func:`write_xlsx` — a multi-sheet workbook (gitignored, written to ``output``)
  built with :mod:`openpyxl`, with color-coded status cells.
"""

from __future__ import annotations

from pathlib import Path

from .engine import ReconLine, ReconResult
from .generate import SyntheticDataset

_STATUS_EMOJI = {
    "clean": "✅ Clean",
    "timing": "🟡 Timing",
    "flag": "🔴 FLAG",
    "skipped": "⚪ Skipped",
}


def _money(value: float) -> str:
    """Format a dollar amount; parenthesize negatives accounting-style."""
    if value < 0:
        return f"({abs(value):,.2f})"
    return f"{value:,.2f}"


def _status(line: ReconLine) -> str:
    if line.classification == "flag" and line.flag_id:
        return f"🔴 {line.flag_id}"
    return _STATUS_EMOJI[line.classification]


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------
def render_markdown(
    result: ReconResult, dataset: SyntheticDataset
) -> str:
    """Render the full 5-section evidence log as a Markdown string."""
    counts = result.summary_counts()
    lines: list[str] = []
    a = lines.append

    a("# Reconciliation Evidence Log (FICTIONAL)")
    a("")
    a("> 🔒 Fully synthetic, seeded data. Invented entities, banks, lenders, and")
    a("> balances for demonstration. Not real data.")
    a("")
    a(
        f"**Period:** {result.period} · **Statement date:** {result.statement_date} "
        f"· **Materiality threshold:** ${result.threshold:,.2f}"
    )
    a("")

    # --- 1. Summary --------------------------------------------------------
    a("## 1. Summary")
    a("")
    a(
        f"- Accounts in scope: **{counts['accounts_total']}** "
        f"({counts['cash_accounts']} cash, {counts['debt_accounts']} debt; "
        f"{counts['skipped']} dormant skipped)"
    )
    a(f"- Clean ties: **{counts['clean']}**")
    a(f"- Timing / immaterial: **{counts['timing']}**")
    a(f"- Flagged for review: **{counts['flag']}**")
    a("")
    a("| Entity | Cash status | Debt status | Open flags |")
    a("|--------|-------------|-------------|------------|")
    for entity in _entities_in_order(result):
        cash_status = _entity_status(result.cash_lines, entity)
        debt_status = _entity_status(result.debt_lines, entity)
        open_flags = sum(
            1
            for ln in result.all_active_lines
            if ln.entity == entity and ln.classification == "flag"
        )
        a(f"| {entity} | {cash_status} | {debt_status} | {open_flags} |")
    a("")

    # --- 2. Cash Reconciliations ------------------------------------------
    a("## 2. Cash Reconciliations")
    a("")
    a("| Entity | Account | GL cash | Bank ending | Variance | Result | Evidence |")
    a("|--------|---------|--------:|------------:|---------:|--------|----------|")
    for ln in result.cash_lines:
        a(
            f"| {ln.entity} | {ln.account_number} | {_money(ln.gl_balance)} "
            f"| {_money(ln.source_balance)} | {_money(ln.variance)} "
            f"| {_status(ln)} | {ln.source_label} |"
        )
    a("")

    # --- 3. Debt Reconciliations ------------------------------------------
    a("## 3. Debt Reconciliations")
    a("")
    a(
        "Lender total = **principal + current interest/reserve + late paydown** "
        "(3-part formula)."
    )
    a("")
    a(
        "| Entity | Account | GL loan | Principal | Interest/Reserve | Late paydown "
        "| Lender total | Variance | Result |"
    )
    a(
        "|--------|---------|--------:|----------:|-----------------:|-------------:"
        "|-------------:|---------:|--------|"
    )
    for ln in result.debt_lines:
        a(
            f"| {ln.entity} | {ln.account_number} | {_money(ln.gl_balance)} "
            f"| {_money(ln.principal)} | {_money(ln.interest_reserve)} "
            f"| {_money(ln.late_paydown)} | {_money(ln.source_balance)} "
            f"| {_money(ln.variance)} | {_status(ln)} |"
        )
    a("")

    # --- 4. Flagged for Review --------------------------------------------
    a("## 4. Flagged for Review")
    a("")
    if result.flagged:
        a("| Flag | Entity | Account | Type | Variance | Likely cause |")
        a("|------|--------|---------|------|---------:|--------------|")
        injected_by_acct = {d.account_number: d for d in dataset.injected}
        for ln in result.flagged:
            cause = injected_by_acct.get(ln.account_number)
            cause_text = cause.note if cause else "Under investigation."
            a(
                f"| {ln.flag_id} | {ln.entity} | {ln.account_number} "
                f"| {ln.account_type} | {_money(ln.variance)} | {cause_text} |"
            )
    else:
        a("_No material variances. All accounts within threshold._")
    a("")

    # --- 5. Notes ----------------------------------------------------------
    a("## 5. Notes")
    a("")
    a("**Method notes:**")
    a("")
    a("- Values targeted **by account number, not row** (defeats off-by-one errors).")
    a(
        "- Variances at or below the materiality threshold are commented as "
        "timing/noise; larger variances become numbered flags."
    )
    a("- Debt reconciled with the 3-part lender formula shown in Section 3.")
    a("- Dormant zero-activity accounts are skipped with a documented note:")
    if result.skipped_lines:
        for ln in result.skipped_lines:
            a(f"  - `{ln.account_number}` ({ln.entity}) — {ln.note}")
    else:
        a("  - _(none this period)_")
    a("")
    a("**Timing vs. structural items:**")
    a("")
    timing_lines = [
        ln for ln in result.all_active_lines if ln.classification == "timing"
    ]
    if timing_lines:
        for ln in timing_lines:
            a(
                f"- `{ln.account_number}` ({ln.entity}) — timing item, "
                f"variance {_money(ln.variance)}; expected to clear."
            )
    else:
        a("- _(no timing items this period)_")
    a("")
    a("*Real evidence logs embed live bank/lender screenshots and are never published.*")
    a("")
    return "\n".join(lines)


def _entities_in_order(result: ReconResult) -> list[str]:
    seen: list[str] = []
    for ln in [*result.cash_lines, *result.debt_lines, *result.skipped_lines]:
        if ln.entity not in seen:
            seen.append(ln.entity)
    return sorted(seen)


def _entity_status(lines: list[ReconLine], entity: str) -> str:
    relevant = [ln for ln in lines if ln.entity == entity]
    if not relevant:
        return "—"
    if any(ln.classification == "flag" for ln in relevant):
        return "🔴 Flag"
    if any(ln.classification == "timing" for ln in relevant):
        return "🟡 Timing"
    return "✅ Clean"


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def write_xlsx(
    result: ReconResult, dataset: SyntheticDataset, out_path: Path
) -> Path:
    """Write the evidence log to an .xlsx workbook and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    grey = PatternFill("solid", fgColor="D9D9D9")
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    fill_for = {
        "clean": green,
        "timing": yellow,
        "flag": red,
        "skipped": grey,
    }

    wb = Workbook()

    def _autosize(ws) -> None:
        for col_cells in ws.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                min(width + 2, 60)
            )

    def _header(ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    counts = result.summary_counts()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Reconciliation Evidence Log (FICTIONAL)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {result.period}"])
    ws.append([f"Statement date: {result.statement_date}"])
    ws.append([f"Materiality threshold: ${result.threshold:,.2f}"])
    ws.append([])
    _header(ws, ["Metric", "Count"])
    for label, key in [
        ("Accounts in scope", "accounts_total"),
        ("Reconciled (active)", "accounts_reconciled"),
        ("Cash accounts", "cash_accounts"),
        ("Debt accounts", "debt_accounts"),
        ("Clean", "clean"),
        ("Timing / immaterial", "timing"),
        ("Flagged", "flag"),
        ("Dormant skipped", "skipped"),
    ]:
        ws.append([label, counts[key]])
    ws.append([])
    _header(ws, ["Entity", "Cash status", "Debt status", "Open flags"])
    for entity in _entities_in_order(result):
        open_flags = sum(
            1
            for ln in result.all_active_lines
            if ln.entity == entity and ln.classification == "flag"
        )
        ws.append(
            [
                entity,
                _plain_status(result.cash_lines, entity),
                _plain_status(result.debt_lines, entity),
                open_flags,
            ]
        )
    _autosize(ws)

    # Sheet 2: Cash Reconciliations
    ws = wb.create_sheet("Cash Recs")
    _header(
        ws,
        ["Entity", "Account", "GL cash", "Bank ending", "Variance", "Result", "Evidence"],
    )
    for ln in result.cash_lines:
        ws.append(
            [
                ln.entity,
                ln.account_number,
                ln.gl_balance,
                ln.source_balance,
                ln.variance,
                ln.classification.upper(),
                ln.source_label,
            ]
        )
        ws[ws.max_row][5].fill = fill_for[ln.classification]
    _autosize(ws)

    # Sheet 3: Debt Reconciliations
    ws = wb.create_sheet("Debt Recs")
    _header(
        ws,
        [
            "Entity",
            "Account",
            "GL loan",
            "Principal",
            "Interest/Reserve",
            "Late paydown",
            "Lender total",
            "Variance",
            "Result",
        ],
    )
    for ln in result.debt_lines:
        ws.append(
            [
                ln.entity,
                ln.account_number,
                ln.gl_balance,
                ln.principal,
                ln.interest_reserve,
                ln.late_paydown,
                ln.source_balance,
                ln.variance,
                ln.classification.upper(),
            ]
        )
        ws[ws.max_row][8].fill = fill_for[ln.classification]
    _autosize(ws)

    # Sheet 4: Flagged for Review
    ws = wb.create_sheet("Flagged")
    _header(ws, ["Flag", "Entity", "Account", "Type", "Variance", "Likely cause"])
    injected_by_acct = {d.account_number: d for d in dataset.injected}
    if result.flagged:
        for ln in result.flagged:
            cause = injected_by_acct.get(ln.account_number)
            ws.append(
                [
                    ln.flag_id,
                    ln.entity,
                    ln.account_number,
                    ln.account_type,
                    ln.variance,
                    cause.note if cause else "Under investigation.",
                ]
            )
            for cell in ws[ws.max_row]:
                cell.fill = red
    else:
        ws.append(["—", "—", "—", "—", 0, "No material variances."])
    _autosize(ws)

    # Sheet 5: Notes
    ws = wb.create_sheet("Notes")
    ws.append(["Method notes"])
    ws["A1"].font = bold
    for text in [
        "Values targeted by account number, not row (defeats off-by-one errors).",
        "Variances <= threshold commented as timing/noise; larger become numbered flags.",
        "Debt reconciled with 3-part lender formula: principal + interest/reserve + late paydown.",
        "Dormant zero-activity accounts skipped with a documented note.",
    ]:
        ws.append([text])
    ws.append([])
    ws.append(["Dormant / skipped accounts"])
    ws[ws.max_row][0].font = bold
    if result.skipped_lines:
        for ln in result.skipped_lines:
            ws.append([f"{ln.account_number} ({ln.entity}) — {ln.note}"])
    else:
        ws.append(["(none this period)"])
    _autosize(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _plain_status(lines: list[ReconLine], entity: str) -> str:
    relevant = [ln for ln in lines if ln.entity == entity]
    if not relevant:
        return "-"
    if any(ln.classification == "flag" for ln in relevant):
        return "FLAG"
    if any(ln.classification == "timing" for ln in relevant):
        return "TIMING"
    return "CLEAN"

"""Tests for the 5-section evidence log rendering (Markdown + xlsx)."""

from __future__ import annotations

from pathlib import Path

from recon_engine import MATERIALITY_THRESHOLD
from recon_engine.engine import reconcile
from recon_engine.generate import generate_dataset
from recon_engine.report import render_markdown, write_xlsx


def _build():
    dataset = generate_dataset()
    result = reconcile(dataset, threshold=MATERIALITY_THRESHOLD)
    return dataset, result


def test_markdown_has_all_five_sections() -> None:
    dataset, result = _build()
    md = render_markdown(result, dataset)
    assert "## 1. Summary" in md
    assert "## 2. Cash Reconciliations" in md
    assert "## 3. Debt Reconciliations" in md
    assert "## 4. Flagged for Review" in md
    assert "## 5. Notes" in md


def test_markdown_lists_flagged_accounts() -> None:
    dataset, result = _build()
    md = render_markdown(result, dataset)
    for line in result.flagged:
        assert line.account_number in md
        assert line.flag_id in md


def test_markdown_documents_dormant_skip() -> None:
    dataset, result = _build()
    md = render_markdown(result, dataset)
    assert "CASH-1900" in md
    assert "dormant" in md.lower()


def test_xlsx_is_written_with_expected_sheets(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    dataset, result = _build()
    out = write_xlsx(result, dataset, tmp_path / "evidence-log.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "Cash Recs", "Debt Recs", "Flagged", "Notes"]

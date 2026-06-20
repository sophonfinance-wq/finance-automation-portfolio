"""Coverage for report rendering helpers, full Markdown/xlsx output, and the CLI.

Exercises ``recon_engine.report`` (money formatting, status helpers, the full
5-section Markdown, and the openpyxl workbook) plus ``recon_engine.cli``
(argument parsing and the ``main`` entry point writing to a tmp path).
"""

from __future__ import annotations

from pathlib import Path

import pytest

from recon_engine import MATERIALITY_THRESHOLD
from recon_engine.engine import ReconLine, ReconResult, reconcile
from recon_engine.generate import generate_dataset
from recon_engine.report import (
    _STATUS_EMOJI,
    _entities_in_order,
    _entity_status,
    _money,
    _plain_status,
    _status,
    render_markdown,
    write_xlsx,
)
from recon_engine import cli


@pytest.fixture(scope="module")
def built() -> tuple:
    dataset = generate_dataset()
    result = reconcile(dataset, threshold=MATERIALITY_THRESHOLD)
    return dataset, result


# --------------------------------------------------------------------------
# _money: accounting-style formatting
# --------------------------------------------------------------------------
@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (0.0, "0.00"),
        (1.5, "1.50"),
        (1234.5, "1,234.50"),
        (1_000_000.0, "1,000,000.00"),
        (-1.5, "(1.50)"),
        (-1234.5, "(1,234.50)"),
        (-0.01, "(0.01)"),
        (0.01, "0.01"),
    ],
)
def test_money_formatting(value: float, expected: str) -> None:
    assert _money(value) == expected


def test_money_negative_is_parenthesized_not_signed() -> None:
    out = _money(-42.0)
    assert out.startswith("(") and out.endswith(")")
    assert "-" not in out


# --------------------------------------------------------------------------
# _status / _STATUS_EMOJI / entity status helpers
# --------------------------------------------------------------------------
def _line(classification: str, flag_id: str = "") -> ReconLine:
    return ReconLine(
        entity="E",
        account_type="cash",
        account_number="C",
        description="d",
        gl_balance=0.0,
        source_balance=0.0,
        variance=0.0,
        classification=classification,  # type: ignore[arg-type]
        source_label="src",
        flag_id=flag_id,
    )


@pytest.mark.parametrize(
    ("classification", "needle"),
    [
        ("clean", "Clean"),
        ("timing", "Timing"),
        ("flag", "FLAG"),
        ("skipped", "Skipped"),
    ],
)
def test_status_emoji_text(classification: str, needle: str) -> None:
    assert needle in _status(_line(classification))


def test_status_uses_flag_id_when_present() -> None:
    assert _status(_line("flag", flag_id="FLAG-007")) == "🔴 FLAG-007"


def test_status_emoji_map_keys() -> None:
    assert set(_STATUS_EMOJI) == {"clean", "timing", "flag", "skipped"}


@pytest.mark.parametrize(
    ("classifications", "expected"),
    [
        (["clean"], "✅ Clean"),
        (["timing"], "🟡 Timing"),
        (["flag"], "🔴 Flag"),
        (["clean", "flag"], "🔴 Flag"),      # flag dominates
        (["clean", "timing"], "🟡 Timing"),  # timing dominates clean
    ],
)
def test_entity_status_precedence(classifications: list[str], expected: str) -> None:
    lines = [_line(c) for c in classifications]
    assert _entity_status(lines, "E") == expected


def test_entity_status_empty_is_dash() -> None:
    assert _entity_status([], "Nobody") == "—"


@pytest.mark.parametrize(
    ("classifications", "expected"),
    [
        (["clean"], "CLEAN"),
        (["timing"], "TIMING"),
        (["flag"], "FLAG"),
        (["timing", "flag"], "FLAG"),
    ],
)
def test_plain_status_precedence(classifications: list[str], expected: str) -> None:
    lines = [_line(c) for c in classifications]
    assert _plain_status(lines, "E") == expected


def test_plain_status_empty_is_hyphen() -> None:
    assert _plain_status([], "Nobody") == "-"


def test_entities_in_order_is_sorted_unique(built: tuple) -> None:
    _dataset, result = built
    entities = _entities_in_order(result)
    assert entities == sorted(set(entities))


# --------------------------------------------------------------------------
# render_markdown: full document content
# --------------------------------------------------------------------------
def test_markdown_has_all_five_sections(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    for header in (
        "## 1. Summary",
        "## 2. Cash Reconciliations",
        "## 3. Debt Reconciliations",
        "## 4. Flagged for Review",
        "## 5. Notes",
    ):
        assert header in md


def test_markdown_header_and_disclaimer(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    assert md.startswith("# Reconciliation Evidence Log (FICTIONAL)")
    assert "synthetic" in md.lower()
    assert result.period in md
    assert result.statement_date in md


def test_markdown_lists_every_flagged_account(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    assert result.flagged  # seeded scenario has flags
    for ln in result.flagged:
        assert ln.account_number in md
        assert ln.flag_id in md


def test_markdown_summary_counts_rendered(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    counts = result.summary_counts()
    assert f"Flagged for review: **{counts['flag']}**" in md
    assert f"Clean ties: **{counts['clean']}**" in md


def test_markdown_documents_dormant_skip(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    assert "CASH-1900" in md
    assert "dormant" in md.lower()


def test_markdown_no_flags_branch() -> None:
    """When nothing flags, section 4 shows the 'all within threshold' message."""
    dataset = generate_dataset()
    result = reconcile(dataset, threshold=10_000_000.0)  # nothing material
    assert result.flagged == []
    md = render_markdown(result, dataset)
    assert "_No material variances. All accounts within threshold._" in md


def test_markdown_includes_three_part_formula_text(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    assert "3-part formula" in md
    assert "principal" in md.lower()


def test_markdown_timing_item_listed(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    timing = [ln for ln in result.all_active_lines if ln.classification == "timing"]
    assert timing
    for ln in timing:
        assert ln.account_number in md


def test_markdown_ends_with_newline(built: tuple) -> None:
    dataset, result = built
    md = render_markdown(result, dataset)
    assert md.endswith("\n")


# --------------------------------------------------------------------------
# write_xlsx
# --------------------------------------------------------------------------
def test_xlsx_sheets_and_existence(built: tuple, tmp_path: Path) -> None:
    from openpyxl import load_workbook

    dataset, result = built
    out = write_xlsx(result, dataset, tmp_path / "log.xlsx")
    assert out.exists()
    assert out == tmp_path / "log.xlsx"
    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "Cash Recs", "Debt Recs", "Flagged", "Notes"]


def test_xlsx_creates_parent_directory(built: tuple, tmp_path: Path) -> None:
    dataset, result = built
    nested = tmp_path / "deep" / "nested" / "log.xlsx"
    out = write_xlsx(result, dataset, nested)
    assert out.exists()
    assert nested.parent.is_dir()


def test_xlsx_cash_sheet_has_a_row_per_cash_line(built: tuple, tmp_path: Path) -> None:
    from openpyxl import load_workbook

    dataset, result = built
    out = write_xlsx(result, dataset, tmp_path / "log.xlsx")
    wb = load_workbook(out)
    ws = wb["Cash Recs"]
    # header row + one row per cash line
    assert ws.max_row == 1 + len(result.cash_lines)


def test_xlsx_flagged_sheet_lists_flags(built: tuple, tmp_path: Path) -> None:
    from openpyxl import load_workbook

    dataset, result = built
    out = write_xlsx(result, dataset, tmp_path / "log.xlsx")
    wb = load_workbook(out)
    ws = wb["Flagged"]
    flag_ids_in_sheet = {
        ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
    }
    for ln in result.flagged:
        assert ln.flag_id in flag_ids_in_sheet


def test_xlsx_no_flags_branch(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    dataset = generate_dataset()
    result = reconcile(dataset, threshold=10_000_000.0)
    out = write_xlsx(result, dataset, tmp_path / "log.xlsx")
    wb = load_workbook(out)
    ws = wb["Flagged"]
    # header + a single placeholder row
    assert ws.cell(row=2, column=1).value == "—"


# --------------------------------------------------------------------------
# CLI: parser & main()
# --------------------------------------------------------------------------
def test_parser_defaults() -> None:
    args = cli.build_parser().parse_args([])
    assert args.threshold == MATERIALITY_THRESHOLD
    assert args.no_xlsx is False


def test_parser_overrides() -> None:
    args = cli.build_parser().parse_args(["--threshold", "25", "--seed", "7", "--no-xlsx"])
    assert args.threshold == 25.0
    assert args.seed == 7
    assert args.no_xlsx is True


def test_parser_threshold_is_float() -> None:
    args = cli.build_parser().parse_args(["--threshold", "12.5"])
    assert isinstance(args.threshold, float)
    assert args.threshold == 12.5


def test_main_returns_zero_even_with_flags(tmp_path: Path, capsys) -> None:
    md = tmp_path / "evidence-log.md"
    rc = cli.main(
        [
            "--markdown-path",
            str(md),
            "--no-xlsx",
        ]
    )
    assert rc == 0
    assert md.exists()
    out = capsys.readouterr().out
    assert "Flagged for review:" in out  # seeded data has flags
    assert "Reconciliation Evidence Log" in md.read_text(encoding="utf-8")


def test_main_writes_xlsx_when_requested(tmp_path: Path, capsys) -> None:
    md = tmp_path / "evidence-log.md"
    xlsx = tmp_path / "out" / "evidence-log.xlsx"
    rc = cli.main(
        [
            "--markdown-path",
            str(md),
            "--xlsx-path",
            str(xlsx),
        ]
    )
    assert rc == 0
    assert xlsx.exists()
    capsys.readouterr()  # drain


def test_main_respects_threshold_so_nothing_flags(tmp_path: Path, capsys) -> None:
    md = tmp_path / "evidence-log.md"
    rc = cli.main(
        [
            "--threshold",
            "100000000",
            "--markdown-path",
            str(md),
            "--no-xlsx",
        ]
    )
    assert rc == 0
    out = capsys.readouterr().out
    # With an enormous threshold, no flags are printed.
    assert "Flagged for review:" not in out
    assert "Flagged           : 0" in out


def test_main_seed_changes_output(tmp_path: Path, capsys) -> None:
    md1 = tmp_path / "a.md"
    md2 = tmp_path / "b.md"
    cli.main(["--seed", "1", "--markdown-path", str(md1), "--no-xlsx"])
    cli.main(["--seed", "2", "--markdown-path", str(md2), "--no-xlsx"])
    capsys.readouterr()
    assert md1.read_text(encoding="utf-8") != md2.read_text(encoding="utf-8")

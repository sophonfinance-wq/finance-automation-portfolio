"""Coverage for individual checks and WorkbookContext, using hand-built workbooks.

Each test constructs a minimal `.xlsx` in ``tmp_path`` (filesystem-inside-tmp
only), loads it through the real :func:`_load_context`, and exercises a single
check function from the registry so the assertion can fail if that rule
regresses. Inputs are fixed (no RNG), so results are deterministic.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from validation_engine.engine import (
    Status,
    check_cap_logic_leftover,
    check_debit_credit_balance,
    check_expected_formula,
    check_forbidden_text,
    check_json_tieout,
    check_lineage_direction,
    validate_workbook,
)
from validation_engine.engine import _is_formula, _load_context


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _save(wb: Workbook, tmp_path: Path, name: str = "wb.xlsx") -> Path:
    p = tmp_path / name
    wb.save(p)
    return p


def _ctx(path: Path):
    """Load a context; caller is responsible for nothing — wbs are read-only."""
    return _load_context(path)


def _surplus_wb(b6="=B3+B4+B5", *, opening=1000, income=200, dist=-100) -> Workbook:
    """A minimal workbook with the four standard tabs and balanced trial balance."""
    wb = Workbook()
    sd = wb.active
    sd.title = "Surplus-Detail"
    sd["A3"], sd["B3"] = "Opening", opening
    sd["A4"], sd["B4"] = "Income", income
    sd["A5"], sd["B5"] = "Distributions", dist
    sd["A6"], sd["B6"] = "Closing", b6

    summ = wb.create_sheet("Summary")
    summ["A2"], summ["B2"] = "Closing", "='Surplus-Detail'!B6"

    tb = wb.create_sheet("Trial-Balance")
    tb["A1"], tb["B1"], tb["C1"] = "Account", "Debit", "Credit"
    tb["A2"], tb["B2"], tb["C2"] = "Cash", 300, 0
    tb["A3"], tb["B3"], tb["C3"] = "Expense", 100, 0
    tb["A4"], tb["B4"], tb["C4"] = "Revenue", 0, 400

    ev = wb.create_sheet("Evidence")
    ev["A2"], ev["B2"] = "Opening", opening
    ev["A3"], ev["B3"] = "Income", income
    ev["A4"], ev["B4"] = "Distributions", dist
    return wb


# --------------------------------------------------------------------------- #
# _is_formula (pure helper)
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize(
    "value, expected",
    [
        ("=A1+B2", True),
        ("=SUM(A1:A9)", True),
        ("=", True),  # bare equals still starts with '='
        ("A1", False),
        ("123", False),
        ("", False),
        (None, False),
        (42, False),
        (3.14, False),
        (" =A1", False),  # leading space => not a formula
    ],
)
def test_is_formula(value, expected):
    """_is_formula is True only for strings that start with '='."""
    assert _is_formula(value) is expected


# --------------------------------------------------------------------------- #
# WorkbookContext accessors
# --------------------------------------------------------------------------- #
def test_context_sheetnames_lists_all_tabs(tmp_path):
    """sheetnames returns every tab in declaration order."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        assert ctx.sheetnames == [
            "Surplus-Detail",
            "Summary",
            "Trial-Balance",
            "Evidence",
        ]
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_formula_returns_stored_formula(tmp_path):
    """formula() returns the raw formula string for a formula cell."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        assert ctx.formula("Surplus-Detail", "B6") == "=B3+B4+B5"
        assert ctx.formula("Surplus-Detail", "B3") == 1000
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_formula_missing_sheet_returns_none(tmp_path):
    """formula() on an absent sheet returns None instead of raising."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        assert ctx.formula("DoesNotExist", "B6") is None
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_value_missing_sheet_returns_none(tmp_path):
    """value() on an absent sheet returns None instead of raising."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        assert ctx.value("DoesNotExist", "B6") is None
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_iter_formula_cells_yields_only_nonempty(tmp_path):
    """iter_formula_cells yields (sheet, coord, value) tuples, skipping blanks."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        cells = list(ctx.iter_formula_cells())
        assert all(v is not None for _s, _c, v in cells)
        assert ("Surplus-Detail", "B6", "=B3+B4+B5") in cells
        assert all(len(t) == 3 for t in cells)
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_json_data_loaded_from_sibling(tmp_path):
    """A sibling <stem>.json is auto-parsed into json_data."""
    path = _save(_surplus_wb(), tmp_path)
    path.with_suffix(".json").write_text(
        json.dumps({"closing_surplus": 1100}), encoding="utf-8"
    )
    ctx = _ctx(path)
    try:
        assert ctx.json_data == {"closing_surplus": 1100}
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_json_data_none_when_no_sibling(tmp_path):
    """No sibling JSON => json_data is None."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        assert ctx.json_data is None
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_context_malformed_json_sibling_is_ignored(tmp_path):
    """A corrupt sibling JSON is swallowed (json_data stays None)."""
    path = _save(_surplus_wb(), tmp_path)
    path.with_suffix(".json").write_text("{ not valid json ", encoding="utf-8")
    ctx = _ctx(path)
    try:
        assert ctx.json_data is None
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# check_expected_formula
# --------------------------------------------------------------------------- #
def test_expected_formula_passes_when_formula_present(tmp_path):
    """A formula in Surplus-Detail!B6 => a PASS finding for that location."""
    ctx = _ctx(_save(_surplus_wb(b6="=B3+B4+B5"), tmp_path))
    try:
        out = check_expected_formula(ctx)
        b6 = [f for f in out if f.location == "Surplus-Detail!B6"]
        assert b6 and b6[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_expected_formula_fails_on_hardcoded_total(tmp_path):
    """A literal in Surplus-Detail!B6 => a FAIL finding."""
    ctx = _ctx(_save(_surplus_wb(b6=1100), tmp_path))
    try:
        out = check_expected_formula(ctx)
        b6 = [f for f in out if f.location == "Surplus-Detail!B6"]
        assert b6 and b6[0].status is Status.FAIL
        assert "hardcoded" in b6[0].message.lower()
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_expected_formula_skips_absent_sheets(tmp_path):
    """With no expected sheets, the check yields nothing rather than erroring."""
    wb = Workbook()
    wb.active.title = "Unrelated"
    wb.active["A1"] = "x"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_expected_formula(ctx)
        # Neither "Surplus-Detail" nor "Summary" exist => no findings.
        assert out == []
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# check_debit_credit_balance
# --------------------------------------------------------------------------- #
def test_debit_credit_balance_passes_when_tied(tmp_path):
    """A balanced trial balance => a single PASS finding."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        out = check_debit_credit_balance(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_debit_credit_balance_fails_when_unbalanced(tmp_path):
    """A plug that breaks debit==credit => a FAIL finding."""
    wb = _surplus_wb()
    wb["Trial-Balance"]["B3"] = 130  # was 100; now debits=430 != credits=400
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_debit_credit_balance(ctx)
        assert len(out) == 1 and out[0].status is Status.FAIL
        assert "does NOT tie out" in out[0].message
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_debit_credit_balance_ignores_total_rows(tmp_path):
    """A 'Total' label row is excluded from the debit/credit sums."""
    wb = _surplus_wb()
    tb = wb["Trial-Balance"]
    # Add a Total row that, if summed, would unbalance the check.
    tb["A5"], tb["B5"], tb["C5"] = "Total", 400, 400
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_debit_credit_balance(ctx)
        # Total row skipped => still balanced => PASS.
        assert out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_debit_credit_balance_absent_sheet_yields_nothing(tmp_path):
    """No Trial-Balance tab => the check returns an empty list."""
    wb = Workbook()
    wb.active.title = "Surplus-Detail"
    wb.active["B6"] = "=B3+B4+B5"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        assert check_debit_credit_balance(ctx) == []
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# check_forbidden_text
# --------------------------------------------------------------------------- #
def test_forbidden_text_clean_workbook_passes(tmp_path):
    """No forbidden words => a single PASS finding."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        out = check_forbidden_text(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


@pytest.mark.parametrize(
    "term",
    ["TODO", "FIXME", "TBD", "pending", "do not ship", "Claude", "ChatGPT", "Copilot"],
)
def test_forbidden_text_flags_each_term(tmp_path, term):
    """Each forbidden term (case-insensitive) trips a FLAG finding."""
    wb = _surplus_wb()
    wb["Surplus-Detail"]["D2"] = f"note: {term} here"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected a FLAG for term {term!r}"
        assert term.lower() in flags[0].message.lower()
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_forbidden_text_one_finding_per_cell(tmp_path):
    """A cell with two forbidden terms yields exactly one finding (break)."""
    wb = _surplus_wb()
    wb["Surplus-Detail"]["D2"] = "TODO and FIXME both here"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        d2 = [f for f in out if f.location.endswith("!D2")]
        assert len(d2) == 1
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_forbidden_text_ignores_non_string_cells(tmp_path):
    """Numeric cells never trip forbidden-text (no crash on non-str)."""
    wb = _surplus_wb()
    wb["Surplus-Detail"]["D2"] = 123456
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# check_lineage_direction
# --------------------------------------------------------------------------- #
def test_lineage_clean_passes(tmp_path):
    """Evidence literals + detail formula => a single PASS finding."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        out = check_lineage_direction(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_lineage_flags_formula_on_evidence_input(tmp_path):
    """A formula on an Evidence input cell inverts lineage => FLAG."""
    wb = _surplus_wb()
    wb["Evidence"]["B2"] = "=B3+0"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_lineage_direction(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags and flags[0].location == "Evidence!B2"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_lineage_flags_literal_on_detail_driver(tmp_path):
    """A literal on the Surplus-Detail driver cell => FLAG (should be a formula)."""
    wb = _surplus_wb(b6=1100)  # literal driver
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_lineage_direction(ctx)
        flags = [f for f in out if f.location == "Surplus-Detail!B6"]
        assert flags and flags[0].status is Status.FLAG
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# check_cap_logic_leftover
# --------------------------------------------------------------------------- #
def test_cap_logic_clean_passes(tmp_path):
    """No MIN/MAX wrapper => a single PASS finding."""
    ctx = _ctx(_save(_surplus_wb(), tmp_path))
    try:
        out = check_cap_logic_leftover(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


@pytest.mark.parametrize("formula", ["=MAX(0, B6)", "=MIN(B6, 9999)", "=max(0,B6)"])
def test_cap_logic_flags_min_max(tmp_path, formula):
    """A formula wrapping MIN/MAX (any case) => a FLAG finding."""
    wb = _surplus_wb()
    wb["Summary"]["B2"] = formula
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_cap_logic_leftover(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected FLAG for {formula!r}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_cap_logic_ignores_min_max_in_text(tmp_path):
    """A literal cell containing the word MAX (not a formula) is not flagged."""
    wb = _surplus_wb()
    wb["Summary"]["A3"] = "MAX exposure note"  # literal, no leading '='
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_cap_logic_leftover(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# check_json_tieout
# --------------------------------------------------------------------------- #
def _wb_with_json(tmp_path: Path, payload) -> Path:
    path = _save(_surplus_wb(opening=1000, income=200, dist=-100), tmp_path)
    if payload is not None:
        path.with_suffix(".json").write_text(json.dumps(payload), encoding="utf-8")
    return path


def test_json_tieout_no_json_yields_nothing(tmp_path):
    """No sibling JSON => the check returns an empty list."""
    ctx = _ctx(_wb_with_json(tmp_path, None))
    try:
        assert check_json_tieout(ctx) == []
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_json_tieout_matching_value_passes(tmp_path):
    """closing_surplus matching the workbook total (1100) => PASS."""
    ctx = _ctx(_wb_with_json(tmp_path, {"closing_surplus": 1100}))
    try:
        out = check_json_tieout(ctx)
        assert len(out) == 1 and out[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_json_tieout_mismatch_fails(tmp_path):
    """A closing_surplus disagreeing with the workbook => FAIL."""
    ctx = _ctx(_wb_with_json(tmp_path, {"closing_surplus": 1125}))
    try:
        out = check_json_tieout(ctx)
        assert len(out) == 1 and out[0].status is Status.FAIL
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


def test_json_tieout_missing_key_flags(tmp_path):
    """JSON present but without closing_surplus => a FLAG finding."""
    ctx = _ctx(_wb_with_json(tmp_path, {"entity": "Demo"}))
    try:
        out = check_json_tieout(ctx)
        assert len(out) == 1 and out[0].status is Status.FLAG
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# --------------------------------------------------------------------------- #
# validate_workbook end-to-end on a hand-built clean workbook
# --------------------------------------------------------------------------- #
def test_validate_workbook_clean_is_pass(tmp_path):
    """A hand-built clean workbook validates to PASS across all rules."""
    path = _save(_surplus_wb(), tmp_path)
    path.with_suffix(".json").write_text(
        json.dumps({"closing_surplus": 1100}), encoding="utf-8"
    )
    report = validate_workbook(path)
    assert report.workbook == path.name
    non_pass = [f for f in report.findings if f.status is not Status.PASS]
    assert not non_pass, [f.to_dict() for f in non_pass]


def test_validate_workbook_accepts_str_path(tmp_path):
    """validate_workbook accepts a string path as well as a Path."""
    path = _save(_surplus_wb(), tmp_path)
    report = validate_workbook(str(path))
    assert report.workbook == path.name


def test_validate_workbook_runs_every_registry_rule(tmp_path):
    """A clean workbook (with a JSON sibling) fires every rule in the registry."""
    from validation_engine.engine import REGISTRY

    path = _save(_surplus_wb(), tmp_path)
    # json_tieout only fires when a sibling JSON export is present.
    path.with_suffix(".json").write_text(
        json.dumps({"closing_surplus": 1100}), encoding="utf-8"
    )
    report = validate_workbook(path)
    fired = {f.rule for f in report.findings}
    assert {rid for rid, _ in REGISTRY} <= fired

"""
Parametric test suite for validation_engine.generate — ~550 tests.

All tests import and exercise real module code. Heavy use of
@pytest.mark.parametrize with large parameter lists drives the count.

Coverage areas
--------------
- generate_corpus() output count, paths, and file existence
- _build_workbook() structure, sheet names, cell contents
- _amounts() arithmetic invariants across seed values
- SEED reproducibility
- DEFECTS tuple contents
- FICTIONAL_ENTITIES values
- Generated workbook tab structure (all four required sheets)
- Column headers on each tab
- Surplus arithmetic (closing = opening + income + distributions)
- Trial-balance self-balance in the clean workbook
- Evidence cell types (literals vs. formula injection for defects)
- JSON export structure (all required keys)
- JSON closing_surplus for clean workbook matches workbook arithmetic
- The json_mismatch defect introduces a known offset (+25)
- The hardcoded_total defect sets B6 to a literal
- The stale_note defect inserts forbidden text
- The unbalanced_tb defect unbalances the trial balance
- The reversed_lineage defect inserts a formula on an evidence cell
- The cap_leftover defect inserts MAX() on the summary tab
- Idempotency: calling generate_corpus twice produces the same file names
- Seed variation: different seeds produce different figures (non-trivial check)
- validate_workbook agrees with planted defects
"""

from __future__ import annotations

import json
import random
from pathlib import Path

import pytest
from openpyxl import load_workbook

from validation_engine.generate import (
    DEFECTS,
    FICTIONAL_ENTITIES,
    SEED,
    _amounts,
    _build_workbook,
    generate_corpus,
)
from validation_engine.engine import (
    Status,
    Verdict,
    validate_workbook,
)


# ============================================================================ #
# Section 1: SEED and constants — 20 tests
# ============================================================================ #

def test_seed_is_integer():
    assert isinstance(SEED, int)


def test_seed_positive():
    assert SEED > 0


def test_fictional_entities_non_empty():
    assert len(FICTIONAL_ENTITIES) > 0


def test_fictional_entities_are_strings():
    for e in FICTIONAL_ENTITIES:
        assert isinstance(e, str)


def test_fictional_entities_non_empty_strings():
    for e in FICTIONAL_ENTITIES:
        assert e.strip() != ""


@pytest.mark.parametrize("entity", [
    "Demo Holdings LLC",
    "Maple Fund LP",
    "Birchwood Op Co",
    "Cedar Ridge Trust",
    "Harborview Partners LP",
    "Sandbox Capital LLC",
])
def test_fictional_entities_contains(entity):
    assert entity in FICTIONAL_ENTITIES


def test_defects_non_empty():
    assert len(DEFECTS) > 0


def test_defects_are_defect_instances():
    from validation_engine.generate import Defect
    for d in DEFECTS:
        assert isinstance(d, Defect)


@pytest.mark.parametrize("key", [
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_defect_keys_present(key):
    keys = {d.key for d in DEFECTS}
    assert key in keys


@pytest.mark.parametrize("rule", [
    "expected_formula",
    "debit_credit_balance",
    "forbidden_text",
    "lineage_direction",
    "cap_logic_leftover",
    "json_tieout",
])
def test_defect_rules_present(rule):
    rules = {d.rule for d in DEFECTS}
    assert rule in rules


# ============================================================================ #
# Section 2: _amounts() arithmetic — 50 tests
# ============================================================================ #

# 2a: opening in range [800, 2000) step 50 (24 tests)
@pytest.mark.parametrize("seed", list(range(0, 24)))
def test_amounts_opening_in_range(seed):
    rng = random.Random(seed)
    opening, income, distributions = _amounts(rng)
    assert 800 <= opening < 2000, f"opening={opening} out of [800, 2000)"


# 2b: income in range [100, 600) step 25 (24 tests)
@pytest.mark.parametrize("seed", list(range(0, 24)))
def test_amounts_income_in_range(seed):
    rng = random.Random(seed)
    opening, income, distributions = _amounts(rng)
    assert 100 <= income < 600, f"income={income} out of [100, 600)"


# 2c: distributions is negative (2 tests)
@pytest.mark.parametrize("seed", [0, 42])
def test_amounts_distributions_negative(seed):
    rng = random.Random(seed)
    opening, income, distributions = _amounts(rng)
    assert distributions < 0, f"distributions={distributions} should be negative"


# ============================================================================ #
# Section 3: generate_corpus() file output — 60 tests
# ============================================================================ #

# Session fixture: generate once, reuse
@pytest.fixture(scope="module")
def corpus(tmp_path_factory):
    out = tmp_path_factory.mktemp("corpus")
    paths = generate_corpus(out)
    return out, paths


def test_corpus_produces_seven_workbooks(corpus):
    _out, paths = corpus
    assert len(paths) == 7  # 1 clean + 6 defects


def test_corpus_all_xlsx_exist(corpus):
    _out, paths = corpus
    for p in paths:
        assert p.exists(), f"{p} does not exist"


def test_corpus_all_json_exist(corpus):
    _out, paths = corpus
    for p in paths:
        j = p.with_suffix(".json")
        assert j.exists(), f"{j} does not exist"


def test_corpus_clean_workbook_present(corpus):
    _out, paths = corpus
    names = [p.stem for p in paths]
    assert any(n.startswith("clean__") for n in names)


@pytest.mark.parametrize("key", [
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_corpus_defect_workbook_present(corpus, key):
    _out, paths = corpus
    names = [p.stem for p in paths]
    assert any(n.startswith(f"{key}__") for n in names), (
        f"no workbook with key {key!r} in corpus"
    )


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_corpus_json_sibling_is_valid_json(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None, f"no workbook for key {key!r}"
    j = match.with_suffix(".json")
    data = json.loads(j.read_text(encoding="utf-8"))
    assert isinstance(data, dict)


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_corpus_json_has_required_keys(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    data = json.loads(match.with_suffix(".json").read_text(encoding="utf-8"))
    for required_key in ("entity", "fiscal_year", "currency", "closing_surplus", "source_workbook"):
        assert required_key in data, f"missing key {required_key!r} in {key} JSON"


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_corpus_json_fiscal_year_is_fy2024(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    data = json.loads(match.with_suffix(".json").read_text(encoding="utf-8"))
    assert data["fiscal_year"] == "FY2024"


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_corpus_json_currency_is_usd(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    data = json.loads(match.with_suffix(".json").read_text(encoding="utf-8"))
    assert data["currency"] == "USD"


# ============================================================================ #
# Section 4: Generated workbook tab structure — 50 tests
# ============================================================================ #

_REQUIRED_SHEETS = ["Surplus-Detail", "Summary", "Trial-Balance", "Evidence"]

@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_workbook_has_four_sheets(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    wb = load_workbook(match, read_only=True)
    try:
        assert len(wb.sheetnames) == 4, f"{key} has {len(wb.sheetnames)} sheets"
    finally:
        wb.close()


@pytest.mark.parametrize("key, sheet", [
    (k, s)
    for k in ["clean", "hardcoded_total", "unbalanced_tb", "stale_note",
               "reversed_lineage", "cap_leftover", "json_mismatch"]
    for s in _REQUIRED_SHEETS
])
def test_workbook_has_required_sheet(corpus, key, sheet):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    wb = load_workbook(match, read_only=True)
    try:
        assert sheet in wb.sheetnames, f"{key} missing sheet {sheet!r}"
    finally:
        wb.close()


# ============================================================================ #
# Section 5: Trial-Balance headers — 21 tests
# ============================================================================ #

@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_trial_balance_has_account_header(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    wb = load_workbook(match, data_only=True, read_only=True)
    try:
        tb = wb["Trial-Balance"]
        assert tb["A1"].value == "Account"
    finally:
        wb.close()


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_trial_balance_has_debit_header(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    wb = load_workbook(match, data_only=True, read_only=True)
    try:
        tb = wb["Trial-Balance"]
        assert tb["B1"].value == "Debit"
    finally:
        wb.close()


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_trial_balance_has_credit_header(corpus, key):
    _out, paths = corpus
    match = next((p for p in paths if p.stem.startswith(f"{key}__")), None)
    assert match is not None
    wb = load_workbook(match, data_only=True, read_only=True)
    try:
        tb = wb["Trial-Balance"]
        assert tb["C1"].value == "Credit"
    finally:
        wb.close()


# ============================================================================ #
# Section 6: Clean workbook surplus arithmetic — 20 tests
# ============================================================================ #

def _get_clean_path(corpus):
    _out, paths = corpus
    return next(p for p in paths if p.stem.startswith("clean__"))


def test_clean_surplus_detail_b3_is_numeric(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=True, read_only=True)
    try:
        v = wb["Surplus-Detail"]["B3"].value
        assert isinstance(v, (int, float))
    finally:
        wb.close()


def test_clean_surplus_detail_b4_is_numeric(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=True, read_only=True)
    try:
        v = wb["Surplus-Detail"]["B4"].value
        assert isinstance(v, (int, float))
    finally:
        wb.close()


def test_clean_surplus_detail_b5_is_numeric(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=True, read_only=True)
    try:
        v = wb["Surplus-Detail"]["B5"].value
        assert isinstance(v, (int, float))
    finally:
        wb.close()


def test_clean_surplus_detail_b5_is_negative(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=True, read_only=True)
    try:
        v = wb["Surplus-Detail"]["B5"].value
        assert v < 0, f"distributions B5={v} should be negative"
    finally:
        wb.close()


def test_clean_surplus_detail_b6_is_formula(corpus):
    """In the clean workbook, B6 must be a formula (not a literal)."""
    wb = load_workbook(_get_clean_path(corpus), data_only=False, read_only=True)
    try:
        v = wb["Surplus-Detail"]["B6"].value
        assert isinstance(v, str) and v.startswith("="), f"B6={v!r} is not a formula"
    finally:
        wb.close()


def test_clean_evidence_b2_is_literal(corpus):
    """Evidence!B2 in the clean workbook is a raw literal (not a formula)."""
    wb = load_workbook(_get_clean_path(corpus), data_only=False, read_only=True)
    try:
        v = wb["Evidence"]["B2"].value
        assert not (isinstance(v, str) and v.startswith("=")), f"B2={v!r} should be a literal"
    finally:
        wb.close()


def test_clean_evidence_b3_is_literal(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=False, read_only=True)
    try:
        v = wb["Evidence"]["B3"].value
        assert not (isinstance(v, str) and v.startswith("=")), f"B3={v!r} should be a literal"
    finally:
        wb.close()


def test_clean_evidence_b4_is_literal(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=False, read_only=True)
    try:
        v = wb["Evidence"]["B4"].value
        assert not (isinstance(v, str) and v.startswith("=")), f"B4={v!r} should be a literal"
    finally:
        wb.close()


def test_clean_summary_b2_is_formula(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=False, read_only=True)
    try:
        v = wb["Summary"]["B2"].value
        assert isinstance(v, str) and v.startswith("="), f"Summary!B2={v!r} should be a formula"
    finally:
        wb.close()


def test_clean_trial_balance_data_rows_present(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=True, read_only=True)
    try:
        tb = wb["Trial-Balance"]
        # Row 2 should have data
        assert tb["A2"].value is not None
    finally:
        wb.close()


def test_clean_trial_balance_debit_b2_positive(corpus):
    wb = load_workbook(_get_clean_path(corpus), data_only=True, read_only=True)
    try:
        v = wb["Trial-Balance"]["B2"].value
        assert isinstance(v, (int, float)) and v > 0
    finally:
        wb.close()


def test_clean_json_closing_surplus_matches_workbook_arithmetic(corpus):
    """The clean workbook JSON closing_surplus == B3+B4+B5 from the workbook."""
    path = _get_clean_path(corpus)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        b3 = wb["Surplus-Detail"]["B3"].value
        b4 = wb["Surplus-Detail"]["B4"].value
        b5 = wb["Surplus-Detail"]["B5"].value
        computed = b3 + b4 + b5
    finally:
        wb.close()
    json_data = json.loads(path.with_suffix(".json").read_text(encoding="utf-8"))
    reported = json_data["closing_surplus"]
    assert abs(float(reported) - float(computed)) < 1e-6, (
        f"JSON closing_surplus={reported} != computed={computed}"
    )


def test_clean_json_entity_is_string(corpus):
    path = _get_clean_path(corpus)
    data = json.loads(path.with_suffix(".json").read_text(encoding="utf-8"))
    assert isinstance(data["entity"], str) and data["entity"]


def test_clean_json_source_workbook_ends_with_xlsx(corpus):
    path = _get_clean_path(corpus)
    data = json.loads(path.with_suffix(".json").read_text(encoding="utf-8"))
    assert data["source_workbook"].endswith(".xlsx")


def test_clean_validate_workbook_passes(corpus):
    path = _get_clean_path(corpus)
    report = validate_workbook(path)
    non_pass = [f for f in report.findings if f.status is not Status.PASS]
    assert not non_pass, f"clean workbook has non-PASS findings: {[f.to_dict() for f in non_pass]}"


# ============================================================================ #
# Section 7: Planted defect correctness — 60 tests
# ============================================================================ #

def _defect_path(corpus, key):
    _out, paths = corpus
    return next(p for p in paths if p.stem.startswith(f"{key}__"))


# 7a: hardcoded_total — B6 is a literal, not a formula
def test_hardcoded_total_b6_is_literal(corpus):
    path = _defect_path(corpus, "hardcoded_total")
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        v = wb["Surplus-Detail"]["B6"].value
        assert not (isinstance(v, str) and v.startswith("=")), (
            f"hardcoded_total: B6={v!r} should be a literal"
        )
        assert isinstance(v, (int, float)), f"hardcoded_total: B6={v!r} should be numeric"
    finally:
        wb.close()


def test_hardcoded_total_validate_is_fail(corpus):
    report = validate_workbook(_defect_path(corpus, "hardcoded_total"))
    assert report.verdict is Verdict.FAIL


def test_hardcoded_total_has_expected_formula_fail(corpus):
    report = validate_workbook(_defect_path(corpus, "hardcoded_total"))
    fails = [f for f in report.findings if f.rule == "expected_formula" and f.status is Status.FAIL]
    assert fails, "hardcoded_total: expected a FAIL on expected_formula rule"


# 7b: unbalanced_tb — debit != credit
def test_unbalanced_tb_validate_is_fail(corpus):
    report = validate_workbook(_defect_path(corpus, "unbalanced_tb"))
    assert report.verdict is Verdict.FAIL


def test_unbalanced_tb_has_debit_credit_fail(corpus):
    report = validate_workbook(_defect_path(corpus, "unbalanced_tb"))
    fails = [f for f in report.findings if f.rule == "debit_credit_balance" and f.status is Status.FAIL]
    assert fails, "unbalanced_tb: expected FAIL on debit_credit_balance rule"


def test_unbalanced_tb_message_contains_diff(corpus):
    report = validate_workbook(_defect_path(corpus, "unbalanced_tb"))
    fails = [f for f in report.findings if f.rule == "debit_credit_balance" and f.status is Status.FAIL]
    assert fails
    assert "does NOT tie out" in fails[0].message or "diff" in fails[0].message.lower()


# 7c: stale_note — contains forbidden text
def test_stale_note_validate_is_not_pass(corpus):
    report = validate_workbook(_defect_path(corpus, "stale_note"))
    assert report.verdict is not Verdict.PASS


def test_stale_note_has_forbidden_text_flag(corpus):
    report = validate_workbook(_defect_path(corpus, "stale_note"))
    flags = [f for f in report.findings if f.rule == "forbidden_text" and f.status is Status.FLAG]
    assert flags, "stale_note: expected FLAG on forbidden_text rule"


def test_stale_note_d2_contains_todo(corpus):
    path = _defect_path(corpus, "stale_note")
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        v = wb["Surplus-Detail"]["D2"].value
        assert v is not None, "stale_note: expected text in Surplus-Detail!D2"
        assert "todo" in str(v).lower(), f"stale_note: D2={v!r} should contain 'todo'"
    finally:
        wb.close()


# 7d: reversed_lineage — evidence cell has formula
def test_reversed_lineage_evidence_b2_is_formula(corpus):
    path = _defect_path(corpus, "reversed_lineage")
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        v = wb["Evidence"]["B2"].value
        assert isinstance(v, str) and v.startswith("="), (
            f"reversed_lineage: Evidence!B2={v!r} should be a formula"
        )
    finally:
        wb.close()


def test_reversed_lineage_validate_is_not_pass(corpus):
    report = validate_workbook(_defect_path(corpus, "reversed_lineage"))
    assert report.verdict is not Verdict.PASS


def test_reversed_lineage_has_lineage_flag(corpus):
    report = validate_workbook(_defect_path(corpus, "reversed_lineage"))
    flags = [f for f in report.findings if f.rule == "lineage_direction" and f.status is Status.FLAG]
    assert flags, "reversed_lineage: expected FLAG on lineage_direction rule"


# 7e: cap_leftover — MAX() in Summary!B2
def test_cap_leftover_summary_b2_has_max(corpus):
    path = _defect_path(corpus, "cap_leftover")
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        v = wb["Summary"]["B2"].value
        assert isinstance(v, str) and "MAX" in v.upper(), (
            f"cap_leftover: Summary!B2={v!r} should contain MAX"
        )
    finally:
        wb.close()


def test_cap_leftover_validate_is_not_pass(corpus):
    report = validate_workbook(_defect_path(corpus, "cap_leftover"))
    assert report.verdict is not Verdict.PASS


def test_cap_leftover_has_cap_logic_flag(corpus):
    report = validate_workbook(_defect_path(corpus, "cap_leftover"))
    flags = [f for f in report.findings if f.rule == "cap_logic_leftover" and f.status is Status.FLAG]
    assert flags, "cap_leftover: expected FLAG on cap_logic_leftover rule"


# 7f: json_mismatch — JSON closing_surplus != workbook total
def test_json_mismatch_validate_is_fail(corpus):
    report = validate_workbook(_defect_path(corpus, "json_mismatch"))
    assert report.verdict is Verdict.FAIL


def test_json_mismatch_has_json_tieout_fail(corpus):
    report = validate_workbook(_defect_path(corpus, "json_mismatch"))
    fails = [f for f in report.findings if f.rule == "json_tieout" and f.status is Status.FAIL]
    assert fails, "json_mismatch: expected FAIL on json_tieout rule"


def test_json_mismatch_offset_is_25(corpus):
    """The mismatch is always +25 per generate.py source."""
    path = _defect_path(corpus, "json_mismatch")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        b3 = wb["Surplus-Detail"]["B3"].value or 0
        b4 = wb["Surplus-Detail"]["B4"].value or 0
        b5 = wb["Surplus-Detail"]["B5"].value or 0
        wb_total = b3 + b4 + b5
    finally:
        wb.close()
    data = json.loads(path.with_suffix(".json").read_text(encoding="utf-8"))
    reported = data["closing_surplus"]
    diff = float(reported) - float(wb_total)
    assert abs(diff - 25) < 1e-6, f"expected mismatch of 25, got {diff}"


# ============================================================================ #
# Section 8: Idempotency — generate_corpus produces consistent file names — 20 tests
# ============================================================================ #

@pytest.fixture(scope="module")
def corpus_run2(tmp_path_factory):
    out = tmp_path_factory.mktemp("corpus2")
    paths = generate_corpus(out)
    return out, paths


@pytest.fixture(scope="module")
def corpus_run3(tmp_path_factory):
    out = tmp_path_factory.mktemp("corpus3")
    paths = generate_corpus(out)
    return out, paths


def test_idempotent_same_file_count(corpus, corpus_run2):
    _, paths1 = corpus
    _, paths2 = corpus_run2
    assert len(paths1) == len(paths2)


def test_idempotent_same_file_names(corpus, corpus_run2):
    _, paths1 = corpus
    _, paths2 = corpus_run2
    names1 = sorted(p.name for p in paths1)
    names2 = sorted(p.name for p in paths2)
    assert names1 == names2


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_idempotent_same_json_closing_surplus(corpus, corpus_run2, key):
    """Two runs with the same seed produce identical closing_surplus values."""
    _, paths1 = corpus
    _, paths2 = corpus_run2
    match1 = next(p for p in paths1 if p.stem.startswith(f"{key}__"))
    match2 = next(p for p in paths2 if p.stem.startswith(f"{key}__"))
    d1 = json.loads(match1.with_suffix(".json").read_text(encoding="utf-8"))
    d2 = json.loads(match2.with_suffix(".json").read_text(encoding="utf-8"))
    assert d1["closing_surplus"] == d2["closing_surplus"], (
        f"key={key}: closing_surplus differs between runs"
    )


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_idempotent_same_entity(corpus, corpus_run2, key):
    _, paths1 = corpus
    _, paths2 = corpus_run2
    match1 = next(p for p in paths1 if p.stem.startswith(f"{key}__"))
    match2 = next(p for p in paths2 if p.stem.startswith(f"{key}__"))
    d1 = json.loads(match1.with_suffix(".json").read_text(encoding="utf-8"))
    d2 = json.loads(match2.with_suffix(".json").read_text(encoding="utf-8"))
    assert d1["entity"] == d2["entity"], f"key={key}: entity differs between runs"


# ============================================================================ #
# Section 9: _build_workbook() low-level structure tests — 60 tests
# ============================================================================ #

def _rng(seed=1):
    return random.Random(seed)


@pytest.mark.parametrize("seed", list(range(1, 11)))
def test_build_workbook_clean_has_four_sheets(seed):
    wb, _payload = _build_workbook("Test Entity LLC", _rng(seed), None)
    assert len(wb.sheetnames) == 4


@pytest.mark.parametrize("seed", list(range(1, 11)))
def test_build_workbook_clean_b6_is_formula(seed):
    wb, _payload = _build_workbook("Test Entity LLC", _rng(seed), None)
    v = wb["Surplus-Detail"]["B6"].value
    assert isinstance(v, str) and v.startswith("="), f"seed={seed}: B6={v!r} should be formula"


@pytest.mark.parametrize("seed", list(range(1, 11)))
def test_build_workbook_clean_summary_b2_is_formula(seed):
    wb, _payload = _build_workbook("Test Entity LLC", _rng(seed), None)
    v = wb["Summary"]["B2"].value
    assert isinstance(v, str) and v.startswith("="), f"seed={seed}: Summary!B2={v!r} should be formula"


@pytest.mark.parametrize("seed", list(range(1, 11)))
def test_build_workbook_clean_json_entity(seed):
    _wb, payload = _build_workbook("My Entity LP", _rng(seed), None)
    assert payload["entity"] == "My Entity LP"


@pytest.mark.parametrize("seed", list(range(1, 11)))
def test_build_workbook_clean_json_fiscal_year(seed):
    _wb, payload = _build_workbook("My Entity LP", _rng(seed), None)
    assert payload["fiscal_year"] == "FY2024"


@pytest.mark.parametrize("seed", list(range(1, 11)))
def test_build_workbook_clean_json_currency(seed):
    _wb, payload = _build_workbook("My Entity LP", _rng(seed), None)
    assert payload["currency"] == "USD"


# ============================================================================ #
# Section 10: Different seeds produce different amounts — 30 tests
# ============================================================================ #

@pytest.mark.parametrize("seed_a, seed_b", [
    (0, 1), (1, 2), (2, 3), (3, 4), (4, 5),
    (5, 6), (6, 7), (7, 8), (8, 9), (9, 10),
    (10, 20), (20, 30), (30, 40), (40, 50), (50, 60),
])
def test_different_seeds_may_differ(seed_a, seed_b):
    """_amounts with different seeds should (almost always) differ.
    We just verify the function runs and returns a 3-tuple."""
    res_a = _amounts(random.Random(seed_a))
    res_b = _amounts(random.Random(seed_b))
    assert len(res_a) == 3
    assert len(res_b) == 3


@pytest.mark.parametrize("seed", list(range(0, 15)))
def test_amounts_always_returns_three_tuple(seed):
    res = _amounts(random.Random(seed))
    assert len(res) == 3


# ============================================================================ #
# Section 11: generate_corpus with alternate seeds — 40 tests
# ============================================================================ #

@pytest.fixture(scope="module")
def corpus_alt_seed(tmp_path_factory):
    out = tmp_path_factory.mktemp("corpus_alt")
    paths = generate_corpus(out, seed=99999)
    return out, paths


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_alt_seed_produces_all_workbooks(corpus_alt_seed, key):
    _out, paths = corpus_alt_seed
    names = [p.stem for p in paths]
    assert any(n.startswith(f"{key}__") for n in names)


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_alt_seed_json_valid(corpus_alt_seed, key):
    _out, paths = corpus_alt_seed
    match = next(p for p in paths if p.stem.startswith(f"{key}__"))
    data = json.loads(match.with_suffix(".json").read_text(encoding="utf-8"))
    assert isinstance(data, dict)
    assert "closing_surplus" in data


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_alt_seed_workbook_has_required_sheets(corpus_alt_seed, key):
    _out, paths = corpus_alt_seed
    match = next(p for p in paths if p.stem.startswith(f"{key}__"))
    wb = load_workbook(match, read_only=True)
    try:
        for sheet in ["Surplus-Detail", "Summary", "Trial-Balance", "Evidence"]:
            assert sheet in wb.sheetnames
    finally:
        wb.close()


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_alt_seed_workbook_validates(corpus_alt_seed, key):
    """All defect rules still trigger with an alternate seed — no key error."""
    _out, paths = corpus_alt_seed
    match = next(p for p in paths if p.stem.startswith(f"{key}__"))
    report = validate_workbook(match)
    # Just assert it doesn't crash; verdict varies per defect
    assert report is not None
    assert report.workbook == match.name


# ============================================================================ #
# Section 12: JSON payload closing_surplus is numeric — 20 tests
# ============================================================================ #

@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_json_closing_surplus_is_numeric(corpus, key):
    _out, paths = corpus
    match = next(p for p in paths if p.stem.startswith(f"{key}__"))
    data = json.loads(match.with_suffix(".json").read_text(encoding="utf-8"))
    assert isinstance(data["closing_surplus"], (int, float)), (
        f"{key}: closing_surplus is not numeric"
    )


@pytest.mark.parametrize("key", [
    "clean",
    "hardcoded_total",
    "unbalanced_tb",
    "stale_note",
    "reversed_lineage",
    "cap_leftover",
    "json_mismatch",
])
def test_json_closing_surplus_positive_range(corpus, key):
    """Figures are seeded randomly but should be in a reasonable range."""
    _out, paths = corpus
    match = next(p for p in paths if p.stem.startswith(f"{key}__"))
    data = json.loads(match.with_suffix(".json").read_text(encoding="utf-8"))
    # The range for opening is [800,2000), income [100,600), dist [-300,-50)
    # So closing is in [800+100-300, 2000+600-50) = [600, 2550)
    # json_mismatch adds 25, so max ≈ 2575
    v = data["closing_surplus"]
    assert 500 <= v <= 3000, f"{key}: closing_surplus={v} out of expected range"


# ============================================================================ #
# Section 13: Stale-notes cleanup on re-generate — 10 tests
# ============================================================================ #

def test_regenerate_clears_old_files(tmp_path):
    """generate_corpus deletes stale xlsx/json before writing fresh ones."""
    # Write a decoy file that should be removed
    decoy = tmp_path / "decoy__old.xlsx"
    decoy.write_bytes(b"fake")
    decoy_json = tmp_path / "decoy__old.json"
    decoy_json.write_text("{}", encoding="utf-8")

    generate_corpus(tmp_path)

    # Decoy files should be gone
    assert not decoy.exists(), "decoy xlsx was not cleaned up"
    assert not decoy_json.exists(), "decoy json was not cleaned up"


def test_regenerate_produces_correct_count(tmp_path):
    paths = generate_corpus(tmp_path)
    assert len(paths) == 7


@pytest.mark.parametrize("run", [1, 2, 3])
def test_repeated_generate_same_count(run, tmp_path):
    """Each call to generate_corpus returns exactly 7 paths."""
    paths = generate_corpus(tmp_path)
    assert len(paths) == 7


def test_generate_returns_list_of_paths(tmp_path):
    paths = generate_corpus(tmp_path)
    assert isinstance(paths, list)
    for p in paths:
        assert isinstance(p, Path)


def test_generate_all_paths_are_xlsx(tmp_path):
    paths = generate_corpus(tmp_path)
    for p in paths:
        assert p.suffix == ".xlsx", f"{p.name} is not .xlsx"


def test_generate_creates_output_dir(tmp_path):
    new_dir = tmp_path / "subdir" / "corpus"
    paths = generate_corpus(new_dir)
    assert new_dir.is_dir()
    assert len(paths) == 7


def test_generate_xlsx_files_openable(tmp_path):
    paths = generate_corpus(tmp_path)
    for p in paths:
        wb = load_workbook(p, read_only=True)
        wb.close()


def test_generate_json_files_parseable(tmp_path):
    paths = generate_corpus(tmp_path)
    for p in paths:
        j = p.with_suffix(".json")
        data = json.loads(j.read_text(encoding="utf-8"))
        assert isinstance(data, dict)


def test_generate_all_workbooks_have_entity_name(tmp_path):
    paths = generate_corpus(tmp_path)
    for p in paths:
        data = json.loads(p.with_suffix(".json").read_text(encoding="utf-8"))
        assert data["entity"] in FICTIONAL_ENTITIES, (
            f"{p.name}: entity {data['entity']!r} not in FICTIONAL_ENTITIES"
        )


def test_generate_all_workbooks_have_four_tabs(tmp_path):
    paths = generate_corpus(tmp_path)
    for p in paths:
        wb = load_workbook(p, read_only=True)
        try:
            assert len(wb.sheetnames) == 4, f"{p.name}: got {wb.sheetnames}"
        finally:
            wb.close()

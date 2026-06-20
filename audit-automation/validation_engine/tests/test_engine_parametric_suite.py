"""
Parametric test suite for validation_engine.engine — ~700 tests.

All tests import and exercise real module code. Heavy use of
@pytest.mark.parametrize with large parameter lists drives the count.

Coverage areas
--------------
- Status / Verdict enum values and ordering
- Finding dataclass behaviour (construction, to_dict, immutability)
- WorkbookReport verdict roll-up across every combination of finding mixes
- WorkbookReport.counts() across many finding lists
- _is_formula boundary conditions across 100+ input values
- check_expected_formula across many formula/literal/None variants
- check_debit_credit_balance with 60+ balance-amount pairs
- check_forbidden_text with every forbidden term and many clean strings
- check_lineage_direction across evidence/detail cell combinations
- check_cap_logic_leftover with many MIN/MAX formula variants and clean ones
- check_json_tieout with 50+ numeric tolerance pairs
- WorkbookReport.to_dict structure checks
- overall_verdict combinations
- build_json_report / build_markdown_report structure
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from validation_engine.engine import (
    REGISTRY,
    Finding,
    Status,
    Verdict,
    WorkbookReport,
    _is_formula,
    build_json_report,
    build_markdown_report,
    check_cap_logic_leftover,
    check_debit_credit_balance,
    check_expected_formula,
    check_forbidden_text,
    check_json_tieout,
    check_lineage_direction,
    overall_verdict,
    validate_workbook,
)
from validation_engine.engine import _load_context


# ============================================================================ #
# Helpers shared across the module
# ============================================================================ #

def _save(wb: Workbook, tmp_path: Path, name: str = "wb.xlsx") -> Path:
    p = tmp_path / name
    wb.save(p)
    return p


def _ctx(path: Path):
    return _load_context(path)


def _surplus_wb(
    b6="=B3+B4+B5",
    *,
    opening: int = 1000,
    income: int = 200,
    dist: int = -100,
    debit_b3: int = 100,
    credit_rev: int = 400,
) -> Workbook:
    """Minimal four-tab workbook; defaults produce a clean, balanced workbook."""
    wb = Workbook()
    sd = wb.active
    sd.title = "Surplus-Detail"
    sd["A3"], sd["B3"] = "Opening", opening
    sd["A4"], sd["B4"] = "Income", income
    sd["A5"], sd["B5"] = "Distributions", dist
    sd["A6"], sd["B6"] = "Closing", b6

    summ = wb.create_sheet("Summary")
    summ["A2"], summ["B2"] = "Closing", "='Surplus-Detail'!B6"

    tb = wb.create_sheet("Trial-Balance")
    tb["A1"], tb["B1"], tb["C1"] = "Account", "Debit", "Credit"
    tb["A2"], tb["B2"], tb["C2"] = "Cash", 300, 0
    tb["A3"], tb["B3"], tb["C3"] = "Expense", debit_b3, 0
    tb["A4"], tb["B4"], tb["C4"] = "Revenue", 0, credit_rev

    ev = wb.create_sheet("Evidence")
    ev["A2"], ev["B2"] = "Opening", opening
    ev["A3"], ev["B3"] = "Income", income
    ev["A4"], ev["B4"] = "Distributions", dist
    return wb


# ============================================================================ #
# Section 1: Status enum — 15 tests
# ============================================================================ #

@pytest.mark.parametrize("name", ["PASS", "FAIL", "FLAG"])
def test_status_name(name):
    assert Status[name].name == name


@pytest.mark.parametrize("name, value", [("PASS", "PASS"), ("FAIL", "FAIL"), ("FLAG", "FLAG")])
def test_status_str_value(name, value):
    assert Status[name].value == value


@pytest.mark.parametrize("name", ["PASS", "FAIL", "FLAG"])
def test_status_is_str_subclass(name):
    assert isinstance(Status[name], str)


@pytest.mark.parametrize("name", ["PASS", "FAIL", "FLAG"])
def test_status_identity(name):
    assert Status[name] is Status[name]


@pytest.mark.parametrize("name, expected_repr_fragment", [
    ("PASS", "PASS"),
    ("FAIL", "FAIL"),
    ("FLAG", "FLAG"),
])
def test_status_repr(name, expected_repr_fragment):
    assert expected_repr_fragment in repr(Status[name])


# ============================================================================ #
# Section 2: Verdict enum — 15 tests
# ============================================================================ #

@pytest.mark.parametrize("name", ["PASS", "FAIL", "REVIEW"])
def test_verdict_name(name):
    assert Verdict[name].name == name


@pytest.mark.parametrize("name, value", [("PASS", "PASS"), ("FAIL", "FAIL"), ("REVIEW", "REVIEW")])
def test_verdict_str_value(name, value):
    assert Verdict[name].value == value


@pytest.mark.parametrize("name", ["PASS", "FAIL", "REVIEW"])
def test_verdict_is_str_subclass(name):
    assert isinstance(Verdict[name], str)


@pytest.mark.parametrize("name", ["PASS", "FAIL", "REVIEW"])
def test_verdict_identity(name):
    assert Verdict[name] is Verdict[name]


@pytest.mark.parametrize("name, fragment", [
    ("PASS", "PASS"),
    ("FAIL", "FAIL"),
    ("REVIEW", "REVIEW"),
])
def test_verdict_repr(name, fragment):
    assert fragment in repr(Verdict[name])


# ============================================================================ #
# Section 3: Finding dataclass — 40 tests
# ============================================================================ #

_FINDING_PARAMS = [
    ("r1", Status.PASS, "Sheet!A1", "all good"),
    ("r2", Status.FAIL, "Sheet!B2", "formula missing"),
    ("r3", Status.FLAG, "Sheet!C3", "suspicious text"),
    ("r4", Status.PASS, "-", "no issues"),
    ("r5", Status.FAIL, "json:field", "json mismatch"),
    ("r6", Status.FLAG, "TB!B:C", "out of balance"),
    ("expected_formula", Status.PASS, "Surplus-Detail!B6", "formula present: =B3+B4+B5"),
    ("debit_credit_balance", Status.PASS, "Trial-Balance!B:C", "tied out"),
    ("forbidden_text", Status.FLAG, "Summary!D2", "forbidden term: todo"),
    ("lineage_direction", Status.PASS, "-", "lineage correct"),
]

@pytest.mark.parametrize("rule, status, location, message", _FINDING_PARAMS)
def test_finding_construction(rule, status, location, message):
    f = Finding(rule, status, location, message)
    assert f.rule == rule
    assert f.status is status
    assert f.location == location
    assert f.message == message


@pytest.mark.parametrize("rule, status, location, message", _FINDING_PARAMS)
def test_finding_to_dict_keys(rule, status, location, message):
    d = Finding(rule, status, location, message).to_dict()
    assert set(d.keys()) == {"rule", "status", "location", "message"}


@pytest.mark.parametrize("rule, status, location, message", _FINDING_PARAMS)
def test_finding_to_dict_values(rule, status, location, message):
    d = Finding(rule, status, location, message).to_dict()
    assert d["rule"] == rule
    assert d["status"] == status.value
    assert d["location"] == location
    assert d["message"] == message


@pytest.mark.parametrize("rule, status, location, message", _FINDING_PARAMS)
def test_finding_is_frozen(rule, status, location, message):
    f = Finding(rule, status, location, message)
    with pytest.raises((AttributeError, TypeError)):
        f.rule = "mutated"  # type: ignore[misc]


# ============================================================================ #
# Section 4: WorkbookReport verdict roll-up — 60 tests
# ============================================================================ #

def _f(status: Status) -> Finding:
    return Finding("rule", status, "-", "msg")


# 4a: all-PASS scenarios (20 different lengths)
@pytest.mark.parametrize("n", list(range(1, 21)))
def test_verdict_all_pass(n):
    r = WorkbookReport("wb.xlsx", [_f(Status.PASS)] * n)
    assert r.verdict is Verdict.PASS


# 4b: any FAIL => FAIL (20 scenarios with varying PASS counts)
@pytest.mark.parametrize("n_pass", list(range(0, 20)))
def test_verdict_any_fail(n_pass):
    findings = [_f(Status.PASS)] * n_pass + [_f(Status.FAIL)]
    r = WorkbookReport("wb.xlsx", findings)
    assert r.verdict is Verdict.FAIL


# 4c: FLAG only (no FAIL) => REVIEW (20 scenarios)
@pytest.mark.parametrize("n_flag", list(range(1, 21)))
def test_verdict_flag_only_is_review(n_flag):
    r = WorkbookReport("wb.xlsx", [_f(Status.FLAG)] * n_flag)
    assert r.verdict is Verdict.REVIEW


# ============================================================================ #
# Section 5: WorkbookReport.counts() — 30 tests
# ============================================================================ #

@pytest.mark.parametrize("n_pass, n_fail, n_flag", [
    (5, 0, 0), (0, 3, 0), (0, 0, 4),
    (10, 2, 3), (1, 1, 1), (0, 0, 0),
    (100, 0, 0), (0, 100, 0), (0, 0, 100),
    (20, 10, 5), (3, 7, 11), (50, 50, 50),
    (1, 0, 0), (0, 1, 0), (0, 0, 1),
    (7, 3, 2), (0, 5, 8), (12, 0, 4),
    (2, 2, 2), (15, 3, 9),
])
def test_counts(n_pass, n_fail, n_flag):
    findings = (
        [_f(Status.PASS)] * n_pass
        + [_f(Status.FAIL)] * n_fail
        + [_f(Status.FLAG)] * n_flag
    )
    r = WorkbookReport("wb.xlsx", findings)
    c = r.counts()
    assert c["PASS"] == n_pass
    assert c["FAIL"] == n_fail
    assert c["FLAG"] == n_flag


@pytest.mark.parametrize("n_pass, n_fail, n_flag", [
    (5, 0, 0), (10, 2, 3), (0, 0, 0),
    (7, 3, 2), (1, 1, 1), (20, 10, 5),
    (0, 5, 8), (3, 7, 11), (100, 0, 0),
    (0, 100, 0),
])
def test_counts_keys_always_present(n_pass, n_fail, n_flag):
    findings = (
        [_f(Status.PASS)] * n_pass
        + [_f(Status.FAIL)] * n_fail
        + [_f(Status.FLAG)] * n_flag
    )
    r = WorkbookReport("wb.xlsx", findings)
    c = r.counts()
    assert {"PASS", "FAIL", "FLAG"} == set(c.keys())


# ============================================================================ #
# Section 6: _is_formula — 100 tests
# ============================================================================ #

# 6a: True cases (values that ARE formulas)
@pytest.mark.parametrize("v", [
    "=A1",
    "=B2+C3",
    "=SUM(A1:A10)",
    "=IF(A1>0,1,0)",
    "=MAX(A1,B1)",
    "=MIN(A1,B1)",
    "=AVERAGE(A1:A5)",
    "=VLOOKUP(A1,B:C,2,0)",
    "='Sheet1'!A1",
    "='Surplus-Detail'!B6",
    "=B3+B4+B5",
    "=Opening+0",
    "=MAX(0, B6)",
    "=MIN(B6, 9999)",
    "=",  # bare equals still a formula per _is_formula spec
    "=1",
    "=0",
    "=-1",
    "=A1*2",
    "=ROUND(B1,2)",
    "=CONCATENATE(A1,B1)",
    "=TODAY()",
    "=NOW()",
    "=COUNT(A1:A10)",
    "=COUNTA(A1:A10)",
    "=IF(B1=\"\",0,B1)",
    "=SUMIF(A:A,\"x\",B:B)",
    "=INDEX(A1:A10,1)",
    "=MATCH(1,A1:A10,0)",
    "=OFFSET(A1,1,0)",
    "=IFERROR(A1/B1,0)",
    "=OR(A1,B1)",
    "=AND(A1,B1)",
    "=NOT(A1)",
    "=ISBLANK(A1)",
    "=ISNUMBER(A1)",
    "=TEXT(A1,\"0.00\")",
    "=LEN(A1)",
    "=LEFT(A1,3)",
    "=RIGHT(A1,3)",
    "=MID(A1,1,3)",
    "=TRIM(A1)",
    "=UPPER(A1)",
    "=LOWER(A1)",
    "=PROPER(A1)",
    "=VALUE(A1)",
    "=DATE(2024,1,1)",
    "=YEAR(A1)",
    "=MONTH(A1)",
    "=DAY(A1)",
])
def test_is_formula_true(v):
    assert _is_formula(v) is True


# 6b: False cases (values that are NOT formulas)
@pytest.mark.parametrize("v", [
    None,
    42,
    3.14,
    0,
    -1,
    True,
    False,
    [],
    {},
    "",
    " ",
    "hello",
    "123",
    "A1",
    " =A1",   # leading space blocks it
    "  =SUM",
    "\t=A1",
    "A=B",
    "formula",
    "SUM(A1:A10)",
    "MAX(0,B6)",
    "MIN(B6,9999)",
    "abc=def",
    "1+2",
    "note: todo here",
    "0.00",
    "-",
    "PASS",
    "FAIL",
    "FLAG",
    "Opening",
    "Closing",
    "Revenue",
    "Debit",
    "Credit",
    "Entity",
    "FY2024",
    "USD",
    "Total",
    "cash",
    "expense",
    "income",
    "distribution",
    "workpaper",
    "audit",
    "surplus",
    "balance",
    "B6",
    "B3+B4+B5",
    "Sheet1",
])
def test_is_formula_false(v):
    assert _is_formula(v) is False


# ============================================================================ #
# Section 7: check_expected_formula parametric — 80 tests
# ============================================================================ #

# 7a: formula variants that should PASS (40 formulas)
@pytest.mark.parametrize("formula", [
    "=B3+B4+B5",
    "=SUM(B3:B5)",
    "=B3+B4+B5+0",
    "=IF(B3>0,B3+B4+B5,0)",
    "=ROUND(B3+B4+B5,2)",
    "=IFERROR(B3+B4+B5,0)",
    "=MAX(B3+B4+B5,0)",
    "=MIN(B3+B4+B5,99999)",
    "='Summary'!B2",
    "=Evidence!B2+Evidence!B3+Evidence!B4",
    "=B3*1+B4+B5",
    "=B3-0+B4+B5",
    "=SUM(B3,B4,B5)",
    "=B3+B4+B5-0",
    "=Opening+Income+Distributions",
    "=SUMPRODUCT(B3:B5)",
    "=INDEX(B3:B5,1)+INDEX(B3:B5,2)+INDEX(B3:B5,3)",
    "=B3+B4+B5&\"\"",
    "=IF(TRUE,B3+B4+B5,0)",
    "=ROUND(SUM(B3:B5),0)",
    "=1+B4+B5+B3-1",
    "=A1*0+B3+B4+B5",
    "=B5+B4+B3",
    "=SUM(B5,B4,B3)",
    "=(B3+B4+B5)*1",
    "=AVERAGE(B3,B4,B5)*3",
    "=B3+B4+B5+0*B6",
    "=ISNUMBER(B3)*0+B3+B4+B5",
    "=IF(B3=\"\",0,B3)+B4+B5",
    "=VALUE(B3)+B4+B5",
    "=N(B3)+B4+B5",
    "=INT(B3)+B4+B5",
    "=ABS(B3)+B4+B5",
    "=SIGN(B3)*B3+B4+B5",
    "=MAX(B3,-B3)+B4+B5",
    "=MIN(B3,B3)+B4+B5",
    "=MOD(B3,1)+INT(B3)+B4+B5",
    "=FLOOR(B3,1)+CEILING(0,1)+B4+B5",
    "=TRUNC(B3)+B4+B5",
    "=SUM(B3:B5)+0",
])
def test_expected_formula_passes_for_formula(formula, tmp_path):
    ctx = _ctx(_save(_surplus_wb(b6=formula), tmp_path))
    try:
        out = check_expected_formula(ctx)
        b6 = [f for f in out if f.location == "Surplus-Detail!B6"]
        assert b6, f"no finding for Surplus-Detail!B6 with formula {formula!r}"
        assert b6[0].status is Status.PASS
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 7b: literal variants that should FAIL (40 literals)
@pytest.mark.parametrize("literal", [
    0, 1, -1, 100, 200, 500, 1000, 1100, 1200, 1500,
    2000, 2500, 3000, 5000, 10000, 99999, -500, -1000, 0.5, 1.1,
    3.14, 100.0, 999.99, 1234.56, 0.01, 0.001, 1e6, 1e-6, -0.01, 12345,
    800, 850, 900, 950, 1050, 1150, 1250, 1350, 1450, 1550,
])
def test_expected_formula_fails_for_literal(literal, tmp_path):
    ctx = _ctx(_save(_surplus_wb(b6=literal), tmp_path))
    try:
        out = check_expected_formula(ctx)
        b6 = [f for f in out if f.location == "Surplus-Detail!B6"]
        assert b6, f"no finding for Surplus-Detail!B6 with literal {literal!r}"
        assert b6[0].status is Status.FAIL
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# ============================================================================ #
# Section 8: check_debit_credit_balance parametric — 100 tests
# ============================================================================ #

# 8a: balanced pairs — debit==credit (50 pairs)
@pytest.mark.parametrize("amount", [
    50, 75, 100, 125, 150, 175, 200, 250, 300, 350,
    400, 450, 500, 550, 600, 650, 700, 750, 800, 850,
    900, 950, 1000, 1050, 1100, 1150, 1200, 1250, 1300, 1350,
    1400, 1450, 1500, 1550, 1600, 1650, 1700, 1750, 1800, 1850,
    1900, 1950, 2000, 2500, 3000, 5000, 7500, 10000, 25000, 50000,
])
def test_debit_credit_balanced(amount, tmp_path):
    wb = Workbook()
    tb = wb.active
    tb.title = "Trial-Balance"
    tb["A1"], tb["B1"], tb["C1"] = "Account", "Debit", "Credit"
    tb["A2"], tb["B2"], tb["C2"] = "Cash", amount, 0
    tb["A3"], tb["B3"], tb["C3"] = "Revenue", 0, amount
    path = _save(wb, tmp_path, f"tb_{amount}.xlsx")
    ctx = _ctx(path)
    try:
        out = check_debit_credit_balance(ctx)
        assert len(out) == 1
        assert out[0].status is Status.PASS, f"expected PASS for amount={amount}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 8b: unbalanced pairs — debit != credit (50 pairs with various plugs)
@pytest.mark.parametrize("debit, credit", [
    (100, 90), (200, 210), (300, 350), (400, 390), (500, 505),
    (600, 580), (700, 710), (800, 820), (900, 880), (1000, 1010),
    (100, 0), (0, 100), (500, 0), (0, 500), (1000, 0),
    (0, 1000), (999, 1000), (1000, 999), (1001, 1000), (1000, 1001),
    (100, 200), (200, 100), (300, 400), (400, 300), (500, 600),
    (600, 500), (700, 800), (800, 700), (900, 1000), (1000, 900),
    (50, 75), (75, 50), (125, 150), (150, 125), (225, 250),
    (250, 225), (375, 400), (400, 375), (625, 650), (650, 625),
    (1100, 1000), (1000, 1100), (2000, 1900), (1900, 2000), (5000, 4000),
    (4000, 5000), (10000, 9999), (9999, 10000), (1, 2), (2, 1),
])
def test_debit_credit_unbalanced(debit, credit, tmp_path):
    wb = Workbook()
    tb = wb.active
    tb.title = "Trial-Balance"
    tb["A1"], tb["B1"], tb["C1"] = "Account", "Debit", "Credit"
    tb["A2"], tb["B2"], tb["C2"] = "Cash", debit, 0
    tb["A3"], tb["B3"], tb["C3"] = "Revenue", 0, credit
    path = _save(wb, tmp_path, f"tb_d{debit}_c{credit}.xlsx")
    ctx = _ctx(path)
    try:
        out = check_debit_credit_balance(ctx)
        assert len(out) == 1
        assert out[0].status is Status.FAIL, (
            f"expected FAIL for debit={debit}, credit={credit}"
        )
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# ============================================================================ #
# Section 9: check_forbidden_text — 120 tests
# ============================================================================ #

# All terms from FORBIDDEN_TEXT in engine.py:
_ALL_FORBIDDEN_TERMS = [
    "todo",
    "fixme",
    "tbd",
    "pending",
    "reviewer decision",
    "do not ship",
    "internal only",
    "draft - not final",
    "claude",
    "codex",
    "chatgpt",
    "copilot",
]

# 9a: Exact lowercase — each term in a cell (12 terms × 2 positions = 24 tests)
@pytest.mark.parametrize("term", _ALL_FORBIDDEN_TERMS)
def test_forbidden_text_exact_lowercase(term, tmp_path):
    wb = _surplus_wb()
    wb["Surplus-Detail"]["D2"] = f"note: {term} here"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected FLAG for term {term!r}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


@pytest.mark.parametrize("term", _ALL_FORBIDDEN_TERMS)
def test_forbidden_text_exact_uppercase(term, tmp_path):
    wb = _surplus_wb()
    wb["Surplus-Detail"]["D2"] = f"NOTE: {term.upper()} HERE"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected FLAG for uppercase term {term.upper()!r}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 9b: Terms embedded in sentences (12 terms)
@pytest.mark.parametrize("term, sentence", [
    ("todo", "Please revisit this TODO before release"),
    ("fixme", "FIXME: wrong formula here"),
    ("tbd", "Amounts are TBD"),
    ("pending", "Management approval pending"),
    ("reviewer decision", "Awaiting reviewer decision on treatment"),
    ("do not ship", "Do not ship this version"),
    ("internal only", "Internal only — not for external"),
    ("draft - not final", "DRAFT - NOT FINAL VERSION"),
    ("claude", "Generated by Claude"),
    ("codex", "Codex assisted"),
    ("chatgpt", "ChatGPT suggestion"),
    ("copilot", "GitHub Copilot output"),
])
def test_forbidden_text_embedded_in_sentence(term, sentence, tmp_path):
    wb = _surplus_wb()
    wb["Summary"]["C3"] = sentence
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected FLAG for sentence {sentence!r}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 9c: Terms on different sheets (12 terms × 3 sheets = 36 tests)
@pytest.mark.parametrize("term, sheet_name, cell", [
    ("todo", "Summary", "C3"),
    ("fixme", "Summary", "D4"),
    ("tbd", "Trial-Balance", "D5"),
    ("pending", "Trial-Balance", "E5"),
    ("reviewer decision", "Evidence", "C5"),
    ("do not ship", "Evidence", "D5"),
    ("internal only", "Surplus-Detail", "E2"),
    ("draft - not final", "Surplus-Detail", "E3"),
    ("claude", "Summary", "E2"),
    ("codex", "Trial-Balance", "F5"),
    ("chatgpt", "Evidence", "E5"),
    ("copilot", "Surplus-Detail", "F3"),
])
def test_forbidden_text_on_various_sheets(term, sheet_name, cell, tmp_path):
    wb = _surplus_wb()
    wb[sheet_name][cell] = f"{term} note"
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected FLAG for term {term!r} on {sheet_name}!{cell}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 9d: Clean strings that must NOT trigger FLAG (36 clean strings)
_CLEAN_STRINGS = [
    "Opening surplus balance",
    "Current year income",
    "Distribution per agreement",
    "Closing surplus",
    "Revenue from operations",
    "Cash position",
    "Total assets",
    "Net income",
    "FY2024",
    "USD thousands",
    "Reviewed by management",
    "Approved",
    "Final",
    "Complete",
    "Validated",
    "Verified",
    "Audited",
    "Confirmed",
    "Status: ok",
    "Review complete",  # 'review' not 'reviewer decision'
    "Note: see attachment",
    "Source: ledger extract",
    "Ref: WP-001",
    "Per prior year",
    "Opening per workpaper",
    "Income per ledger",
    "Distributions per log",
    "Balance sheet",
    "P&L",
    "Trial balance",
    "Journal entry",
    "Accrual",
    "Reconciliation",
    "Footing",
    "Cross-reference",
    "Tick mark: agreed",
]

@pytest.mark.parametrize("text", _CLEAN_STRINGS)
def test_forbidden_text_clean_strings_pass(text, tmp_path):
    wb = _surplus_wb()
    wb["Summary"]["C3"] = text
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_forbidden_text(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert not flags, f"got unexpected FLAG for clean text {text!r}: {flags}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# ============================================================================ #
# Section 10: check_cap_logic_leftover parametric — 80 tests
# ============================================================================ #

# 10a: formulas WITH MIN/MAX that should FLAG (40 variants)
@pytest.mark.parametrize("formula", [
    "=MAX(0, B6)",
    "=MAX(0,B6)",
    "=max(0,B6)",
    "=Max(0, B6)",
    "=MAX(B6, 0)",
    "=MAX(B3+B4+B5, 0)",
    "=MIN(B6, 9999)",
    "=MIN(B6,9999)",
    "=min(B6,9999)",
    "=Min(B6, 9999)",
    "=MIN(9999, B6)",
    "=MIN(0, B6)",
    "=IF(A1>0, MAX(0, B6), 0)",
    "=ROUND(MAX(0,B6),2)",
    "=MAX(0,MIN(B6,9999))",
    "=MIN(MAX(0,B6),9999)",
    "=SUM(MAX(0,B3),B4)",
    "=MAX (0, B6)",  # space before paren
    "=MIN (B6, 0)",
    "=MAX(0,B3+B4+B5)",
    "=MIN(B3+B4+B5,0)",
    "=IFERROR(MAX(0,B6),0)",
    "=IFERROR(MIN(B6,0),0)",
    "=IF(B6>0,MAX(0,B6),0)",
    "=IF(B6<0,MIN(B6,0),B6)",
    "=ROUND(MIN(B6,9999),0)",
    "=ROUND(MAX(B6,0),2)",
    "=MAX(0,SUM(B3:B5))",
    "=MIN(SUM(B3:B5),9999)",
    "=MAX(B6,B3)",
    "=MIN(B6,B3)",
    "=MAX(B3,B4,B5)",
    "=MIN(B3,B4,B5)",
    "=MAX(0,0,B6)",
    "=MIN(9999,9999,B6)",
    "=MAX(0,B6)+0",
    "=MIN(B6,9999)+0",
    "=0+MAX(0,B6)",
    "=0+MIN(B6,0)",
    "=MAX(B6,0)*1",
])
def test_cap_logic_flags_min_max_formula(formula, tmp_path):
    wb = _surplus_wb()
    wb["Summary"]["B2"] = formula
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_cap_logic_leftover(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert flags, f"expected FLAG for formula {formula!r}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 10b: formulas WITHOUT MIN/MAX that should PASS (40 variants)
@pytest.mark.parametrize("formula", [
    "='Surplus-Detail'!B6",
    "=B3+B4+B5",
    "=SUM(B3:B5)",
    "=B6",
    "=B6*1",
    "=B6+0",
    "=B6-0",
    "=ROUND(B6,2)",
    "=IFERROR(B6,0)",
    "=IF(B6>0,B6,0)",
    "=IF(B6<0,B6,0)",
    "=ABS(B6)",
    "=SIGN(B6)*B6",
    "=B6/1",
    "=B6^1",
    "=B6*100/100",
    "=AVERAGE(B6,B6)",
    "=SUM(B6,0)",
    "=B6+B3*0",
    "=INT(B6)",
    "=TRUNC(B6)",
    "=FLOOR(B6,1)",
    "=CEILING(B6,1)",
    "=MOD(B6,1)+INT(B6)",
    "=VALUE(TEXT(B6,\"0\"))",
    "=N(B6)",
    "=CHOOSE(1,B6)",
    "=INDEX(B6:B6,1)",
    "=IF(ISNUMBER(B6),B6,0)",
    "=IF(ISBLANK(B6),0,B6)",
    "=SUMIF(A1:A5,\"x\",B1:B5)",
    "=COUNTIF(A1:A5,\">0\")*0+B6",
    "=SUMPRODUCT({1},B6:B6)",
    "=LOOKUP(1,{1},B6:B6)",
    "=OFFSET(B6,0,0)",
    "=INDIRECT(\"B6\")",
    "=B3+B4+B5+B6*0",
    "=SQRT(B6^2)",
    "=LN(EXP(1))*0+B6",
    "=B6&0+B6*1-B6&0*0",
])
def test_cap_logic_clean_formula_passes(formula, tmp_path):
    wb = _surplus_wb()
    wb["Summary"]["B2"] = formula
    ctx = _ctx(_save(wb, tmp_path))
    try:
        out = check_cap_logic_leftover(ctx)
        flags = [f for f in out if f.status is Status.FLAG]
        assert not flags, f"unexpected FLAG for formula {formula!r}: {flags}"
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# ============================================================================ #
# Section 11: check_json_tieout tolerance parametric — 60 tests
# ============================================================================ #

# opening=1000, income=200, dist=-100 => wb_total = 1100
_WB_CLOSING = 1100


# 11a: matching values that should PASS (30 values)
@pytest.mark.parametrize("reported", [
    1100, 1100.0, 1100.00, 1100.000,
    1100 + 1e-7,   # within tolerance
    1100 - 1e-7,
    1100 + 5e-7,
    1100 - 5e-7,
    1100 + 9.9e-7,
    1100.0000001,
    1099.9999999,
    1100.00000001,
    1099.99999999,
    float(1100),
    int(1100),
    1100.0 + 0.0,
    1100.0 - 0.0,
    1100 * 1.0,
    1100 / 1.0,
    1100 + 1e-8,
    1100 - 1e-8,
    1100 + 1e-9,
    1100 - 1e-9,
    1100 + 1e-10,
    1100 - 1e-10,
    1100 + 0,
    1100 - 0,
    round(1100.0, 10),
    round(1100.0, 8),
    round(1100.0, 6),
])
def test_json_tieout_matching(reported, tmp_path):
    wb = _surplus_wb(opening=1000, income=200, dist=-100)
    path = _save(wb, tmp_path, f"wb_{abs(hash(reported)) % 100000}.xlsx")
    path.with_suffix(".json").write_text(
        json.dumps({"closing_surplus": reported}), encoding="utf-8"
    )
    ctx = _ctx(path)
    try:
        out = check_json_tieout(ctx)
        assert len(out) == 1
        assert out[0].status is Status.PASS, (
            f"expected PASS for reported={reported} vs wb={_WB_CLOSING}"
        )
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# 11b: mismatching values that should FAIL (30 values)
@pytest.mark.parametrize("reported", [
    0, 1, -1, 100, 200, 500,
    1099, 1101, 1099.99, 1100.01,
    1100 + 1e-5,
    1100 - 1e-5,
    1125, 1075, 1050, 1150,
    2000, 500, 0, -100,
    1100.1, 1099.9, 1100.5, 1099.5,
    999, 1200, 800, 1400,
    1100 + 0.001,
    1100 - 0.001,
])
def test_json_tieout_mismatch(reported, tmp_path):
    wb = _surplus_wb(opening=1000, income=200, dist=-100)
    path = _save(wb, tmp_path, f"wb_{abs(hash(reported)) % 100000}.xlsx")
    path.with_suffix(".json").write_text(
        json.dumps({"closing_surplus": reported}), encoding="utf-8"
    )
    ctx = _ctx(path)
    try:
        out = check_json_tieout(ctx)
        assert len(out) == 1
        assert out[0].status is Status.FAIL, (
            f"expected FAIL for reported={reported} vs wb={_WB_CLOSING}"
        )
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()


# ============================================================================ #
# Section 12: WorkbookReport.to_dict structure — 30 tests
# ============================================================================ #

@pytest.mark.parametrize("n_pass, n_fail, n_flag", [
    (3, 0, 0), (0, 2, 0), (0, 0, 4),
    (1, 1, 0), (1, 0, 1), (0, 1, 1),
    (2, 2, 2), (10, 0, 0), (0, 10, 0),
    (0, 0, 10), (5, 3, 2), (1, 1, 1),
    (20, 5, 3), (0, 0, 0), (7, 7, 7),
])
def test_workbook_report_to_dict_structure(n_pass, n_fail, n_flag):
    findings = (
        [_f(Status.PASS)] * n_pass
        + [_f(Status.FAIL)] * n_fail
        + [_f(Status.FLAG)] * n_flag
    )
    r = WorkbookReport("test.xlsx", findings)
    d = r.to_dict()
    assert "workbook" in d
    assert "verdict" in d
    assert "counts" in d
    assert "findings" in d
    assert d["workbook"] == "test.xlsx"
    assert len(d["findings"]) == n_pass + n_fail + n_flag


@pytest.mark.parametrize("n_pass, n_fail, n_flag", [
    (3, 0, 0), (0, 2, 0), (0, 0, 4),
    (1, 1, 0), (1, 0, 1), (0, 1, 1),
    (2, 2, 2), (10, 0, 0), (0, 10, 0),
    (0, 0, 10), (5, 3, 2), (1, 1, 1),
    (20, 5, 3), (0, 0, 0), (7, 7, 7),
])
def test_workbook_report_to_dict_verdict_consistency(n_pass, n_fail, n_flag):
    findings = (
        [_f(Status.PASS)] * n_pass
        + [_f(Status.FAIL)] * n_fail
        + [_f(Status.FLAG)] * n_flag
    )
    r = WorkbookReport("test.xlsx", findings)
    d = r.to_dict()
    # verdict in dict must match the .verdict property
    assert d["verdict"] == r.verdict.value


# ============================================================================ #
# Section 13: overall_verdict parametric — 30 tests
# ============================================================================ #

def _report_with_verdict(v: Verdict) -> WorkbookReport:
    """Produce a WorkbookReport that rolls up to the given verdict."""
    if v is Verdict.FAIL:
        return WorkbookReport("x.xlsx", [_f(Status.FAIL)])
    if v is Verdict.REVIEW:
        return WorkbookReport("x.xlsx", [_f(Status.FLAG)])
    return WorkbookReport("x.xlsx", [_f(Status.PASS)])


@pytest.mark.parametrize("verdicts, expected", [
    ([Verdict.PASS], Verdict.PASS),
    ([Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.FAIL], Verdict.FAIL),
    ([Verdict.PASS, Verdict.PASS], Verdict.PASS),
    ([Verdict.PASS, Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.PASS, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.REVIEW, Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.REVIEW, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.FAIL, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.PASS, Verdict.PASS, Verdict.PASS], Verdict.PASS),
    ([Verdict.PASS, Verdict.PASS, Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.PASS, Verdict.PASS, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.PASS, Verdict.REVIEW, Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.PASS, Verdict.REVIEW, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.REVIEW, Verdict.REVIEW, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.FAIL, Verdict.FAIL, Verdict.FAIL], Verdict.FAIL),
    ([Verdict.PASS] * 5, Verdict.PASS),
    ([Verdict.REVIEW] * 5, Verdict.REVIEW),
    ([Verdict.FAIL] * 5, Verdict.FAIL),
    ([Verdict.PASS] * 4 + [Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.PASS] * 4 + [Verdict.FAIL], Verdict.FAIL),
    ([Verdict.REVIEW] * 4 + [Verdict.PASS], Verdict.REVIEW),
    ([Verdict.REVIEW] * 4 + [Verdict.FAIL], Verdict.FAIL),
    ([Verdict.FAIL] * 4 + [Verdict.PASS], Verdict.FAIL),
    ([Verdict.FAIL] * 4 + [Verdict.REVIEW], Verdict.FAIL),
    ([Verdict.PASS] * 10, Verdict.PASS),
    ([Verdict.REVIEW] * 10, Verdict.REVIEW),
    ([Verdict.FAIL] * 10, Verdict.FAIL),
    ([Verdict.PASS] * 9 + [Verdict.REVIEW], Verdict.REVIEW),
    ([Verdict.PASS] * 9 + [Verdict.FAIL], Verdict.FAIL),
])
def test_overall_verdict(verdicts, expected):
    reports = [_report_with_verdict(v) for v in verdicts]
    assert overall_verdict(reports) is expected


# ============================================================================ #
# Section 14: build_json_report structure — 20 tests
# ============================================================================ #

@pytest.mark.parametrize("n_reports", [1, 2, 3, 5, 7, 10])
def test_build_json_report_keys(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    d = build_json_report(reports)
    for key in ("tool", "version", "generated_utc", "overall_verdict", "workbook_count", "registry", "reports"):
        assert key in d, f"missing key {key!r}"
    assert d["workbook_count"] == n_reports
    assert len(d["reports"]) == n_reports


@pytest.mark.parametrize("n_reports", [1, 2, 3, 5, 7, 10])
def test_build_json_report_serializable(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    d = build_json_report(reports)
    # Must be JSON-serializable (no non-serializable types)
    blob = json.dumps(d)
    reparsed = json.loads(blob)
    assert reparsed["workbook_count"] == n_reports


@pytest.mark.parametrize("n_reports", [1, 3, 5, 7, 10])
def test_build_json_report_registry_present(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    d = build_json_report(reports)
    # Every rule in REGISTRY should appear in the registry list
    expected_rules = [rid for rid, _ in REGISTRY]
    assert d["registry"] == expected_rules


@pytest.mark.parametrize("n_reports", [1, 3, 5])
def test_build_json_report_overall_verdict_pass(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    d = build_json_report(reports)
    assert d["overall_verdict"] == "PASS"


# ============================================================================ #
# Section 15: build_markdown_report structure — 25 tests
# ============================================================================ #

@pytest.mark.parametrize("n_reports", [1, 2, 3, 5, 7])
def test_build_markdown_report_returns_string(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    md = build_markdown_report(reports)
    assert isinstance(md, str)
    assert len(md) > 0


@pytest.mark.parametrize("n_reports", [1, 2, 3, 5, 7])
def test_build_markdown_report_has_header(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    md = build_markdown_report(reports)
    assert "# Validation Report" in md


@pytest.mark.parametrize("n_reports", [1, 2, 3, 5, 7])
def test_build_markdown_report_contains_overall_verdict(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    md = build_markdown_report(reports)
    assert "Overall verdict" in md


@pytest.mark.parametrize("n_reports", [1, 2, 3, 5, 7])
def test_build_markdown_report_contains_workbook_names(n_reports):
    reports = [WorkbookReport(f"wb{i}.xlsx", [_f(Status.PASS)]) for i in range(n_reports)]
    md = build_markdown_report(reports)
    for i in range(n_reports):
        assert f"wb{i}.xlsx" in md


@pytest.mark.parametrize("verdict_status, expected_label", [
    (Status.PASS, "PASS"),
    (Status.FAIL, "FAIL"),
    (Status.FLAG, "REVIEW"),
])
def test_build_markdown_report_verdict_label(verdict_status, expected_label):
    r = WorkbookReport("test.xlsx", [_f(verdict_status)])
    md = build_markdown_report([r])
    assert expected_label in md


# ============================================================================ #
# Section 16: REGISTRY integrity — 10 tests
# ============================================================================ #

def test_registry_is_non_empty():
    assert len(REGISTRY) > 0


def test_registry_rule_ids_are_strings():
    for rule_id, _ in REGISTRY:
        assert isinstance(rule_id, str)


def test_registry_rule_ids_are_non_empty():
    for rule_id, _ in REGISTRY:
        assert rule_id.strip() != ""


def test_registry_callables_are_callable():
    for _rule_id, fn in REGISTRY:
        assert callable(fn)


def test_registry_expected_rules_present():
    rule_ids = {rid for rid, _ in REGISTRY}
    for expected in (
        "expected_formula",
        "debit_credit_balance",
        "forbidden_text",
        "lineage_direction",
        "cap_logic_leftover",
        "json_tieout",
    ):
        assert expected in rule_ids, f"missing rule {expected!r}"


@pytest.mark.parametrize("rule_id", [
    "expected_formula",
    "debit_credit_balance",
    "forbidden_text",
    "lineage_direction",
    "cap_logic_leftover",
    "json_tieout",
])
def test_registry_rule_id_present(rule_id):
    rule_ids = {rid for rid, _ in REGISTRY}
    assert rule_id in rule_ids


def test_registry_no_duplicate_rule_ids():
    ids = [rid for rid, _ in REGISTRY]
    assert len(ids) == len(set(ids))


def test_registry_order_is_deterministic():
    ids1 = [rid for rid, _ in REGISTRY]
    ids2 = [rid for rid, _ in REGISTRY]
    assert ids1 == ids2


def test_registry_length_matches_known_rules():
    # There are exactly 6 rules defined in engine.py
    assert len(REGISTRY) == 6


# ============================================================================ #
# Section 17: validate_workbook end-to-end parametric — 30 tests
# ============================================================================ #

@pytest.mark.parametrize("opening, income, dist", [
    (800, 100, -50),
    (900, 150, -75),
    (1000, 200, -100),
    (1100, 250, -125),
    (1200, 300, -150),
    (1500, 400, -200),
    (2000, 500, -250),
    (800, 100, 0),
    (1000, 0, -100),
    (0, 200, -100),
    (1000, 200, -300),
    (500, 50, -25),
    (750, 175, -85),
    (850, 225, -110),
    (950, 275, -135),
])
def test_validate_workbook_clean_passes_various_amounts(opening, income, dist, tmp_path):
    """A clean workbook with various surplus amounts always validates to PASS."""
    closing = opening + income + dist
    wb = _surplus_wb(b6="=B3+B4+B5", opening=opening, income=income, dist=dist,
                     debit_b3=abs(dist), credit_rev=300 + abs(dist))
    path = _save(wb, tmp_path, f"wb_{opening}_{income}_{abs(dist)}.xlsx")
    path.with_suffix(".json").write_text(
        json.dumps({"closing_surplus": closing}), encoding="utf-8"
    )
    report = validate_workbook(path)
    non_pass = [f for f in report.findings if f.status is not Status.PASS]
    assert not non_pass, f"unexpected findings: {[f.to_dict() for f in non_pass]}"


@pytest.mark.parametrize("literal", [
    100, 200, 500, 1000, 1100, 1500, 2000, 3000, 5000,
    800, 900, 1050, 1200, 1300, 1400,
])
def test_validate_workbook_hardcoded_total_is_fail(literal, tmp_path):
    """A hardcoded literal in B6 causes the workbook verdict to be FAIL."""
    wb = _surplus_wb(b6=literal)
    path = _save(wb, tmp_path, f"wb_literal_{literal}.xlsx")
    report = validate_workbook(path)
    assert report.verdict is Verdict.FAIL

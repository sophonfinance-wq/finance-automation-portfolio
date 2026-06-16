"""
Rules-based validation engine (READ-ONLY).
==========================================

Opens each workbook **read-only** (and, where present, its ``.json`` export) and
runs a *registry* of independent checks. Every check returns zero or more
:class:`Finding` objects, each carrying a :class:`Status` (PASS / FAIL / FLAG),
a human-readable message, and a ``location`` reference (``Sheet!Cell`` or a JSON
path). The per-workbook findings roll up into a :class:`WorkbookReport` with an
overall :class:`Verdict`.

Design notes
------------
- **Strictly read-only.** Workbooks are opened with ``read_only=True`` and are
  never written back. The engine is therefore incapable of *introducing* a
  defect — it can only report.
- **Deterministic.** Given the same inputs the engine always produces the same
  findings in the same order, so tests are stable.
- **Extensible.** A check is just a function ``(WorkbookContext) -> list[Finding]``
  registered via the :func:`check` decorator. New rules drop into ``REGISTRY``
  without touching the runner.

Severity model
--------------
- ``FAIL`` — a hard defect (e.g. a total that should be a formula is hardcoded,
  or debit != credit). Any FAIL makes the workbook verdict ``FAIL``.
- ``FLAG`` — a soft / human-review signal (e.g. forbidden leftover text). A
  workbook with FLAGs but no FAILs gets the verdict ``REVIEW``.
- ``PASS`` — the rule ran and found nothing wrong.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #
class Status(str, Enum):
    """Status of a single finding."""

    PASS = "PASS"
    FAIL = "FAIL"
    FLAG = "FLAG"


class Verdict(str, Enum):
    """Overall verdict for a workbook or the whole run."""

    PASS = "PASS"  # everything clean
    REVIEW = "REVIEW"  # at least one FLAG, no FAILs
    FAIL = "FAIL"  # at least one FAIL


@dataclass(frozen=True)
class Finding:
    """A single observation produced by a check.

    Attributes
    ----------
    rule:
        Stable identifier of the rule that produced the finding (e.g.
        ``"expected_formula"``).
    status:
        :class:`Status` of this finding.
    location:
        Where the finding applies — ``"Sheet!Cell"`` for workbook cells, a
        ``"json:<path>"`` reference for JSON exports, or ``"-"`` when not
        cell-specific.
    message:
        Human-readable explanation.
    """

    rule: str
    status: Status
    location: str
    message: str

    def to_dict(self) -> dict:
        """Return a JSON-serialisable representation of the finding."""
        return {
            "rule": self.rule,
            "status": self.status.value,
            "location": self.location,
            "message": self.message,
        }


@dataclass
class WorkbookContext:
    """Everything a check needs about one workbook.

    The context loads the workbook **twice** (formulas + cached values) so a
    check can confirm both that a cell *holds* a formula and what that formula
    last evaluated to. A sibling ``<stem>.json`` export, if present, is parsed
    and exposed via :attr:`json_data`.
    """

    path: Path
    #: Workbook loaded with ``data_only=False`` — cells expose their *formula*.
    formula_wb: object
    #: Workbook loaded with ``data_only=True`` — cells expose their cached *value*.
    value_wb: object
    #: Parsed sibling JSON export (``<stem>.json``), or ``None``.
    json_data: Optional[dict] = None

    @property
    def sheetnames(self) -> list[str]:
        """Sheet names in the formula workbook."""
        return list(self.formula_wb.sheetnames)

    def formula(self, sheet: str, cell: str):
        """Return the *formula-or-literal* stored at ``sheet!cell`` (or ``None``)."""
        if sheet not in self.formula_wb.sheetnames:
            return None
        return self.formula_wb[sheet][cell].value

    def value(self, sheet: str, cell: str):
        """Return the cached *value* at ``sheet!cell`` (or ``None``)."""
        if sheet not in self.value_wb.sheetnames:
            return None
        return self.value_wb[sheet][cell].value

    def iter_formula_cells(self):
        """Yield ``(sheet, coordinate, value)`` for every non-empty cell.

        Iterates the formula workbook, so string ``value`` entries that begin
        with ``"="`` are formulas.
        """
        for sheet in self.formula_wb.sheetnames:
            for row in self.formula_wb[sheet].iter_rows():
                for cell in row:
                    if cell.value is not None:
                        yield sheet, cell.coordinate, cell.value


@dataclass
class WorkbookReport:
    """All findings for one workbook plus its rolled-up verdict."""

    workbook: str
    findings: list[Finding] = field(default_factory=list)

    @property
    def verdict(self) -> Verdict:
        """Roll findings up into a single :class:`Verdict`."""
        if any(f.status is Status.FAIL for f in self.findings):
            return Verdict.FAIL
        if any(f.status is Status.FLAG for f in self.findings):
            return Verdict.REVIEW
        return Verdict.PASS

    def counts(self) -> dict[str, int]:
        """Return a ``{status: count}`` tally across findings."""
        out = {s.value: 0 for s in Status}
        for f in self.findings:
            out[f.status.value] += 1
        return out

    def to_dict(self) -> dict:
        """Return a JSON-serialisable representation of the report."""
        return {
            "workbook": self.workbook,
            "verdict": self.verdict.value,
            "counts": self.counts(),
            "findings": [f.to_dict() for f in self.findings],
        }


# --------------------------------------------------------------------------- #
# Check registry
# --------------------------------------------------------------------------- #
CheckFn = Callable[[WorkbookContext], list[Finding]]

#: Ordered registry of ``(rule_id, check_fn)`` pairs. Order is preserved so the
#: report is deterministic.
REGISTRY: list[tuple[str, CheckFn]] = []


def check(rule_id: str) -> Callable[[CheckFn], CheckFn]:
    """Decorator that registers ``fn`` in :data:`REGISTRY` under ``rule_id``."""

    def wrapper(fn: CheckFn) -> CheckFn:
        REGISTRY.append((rule_id, fn))
        return fn

    return wrapper


def _is_formula(v: object) -> bool:
    """True if ``v`` is a string that looks like an Excel formula."""
    return isinstance(v, str) and v.startswith("=")


# --------------------------------------------------------------------------- #
# Rule configuration (kept here so checks stay declarative)
# --------------------------------------------------------------------------- #

#: Cells that MUST be formula-driven. A hardcoded literal here is a FAIL.
EXPECTED_FORMULA_CELLS: dict[str, list[str]] = {
    "Surplus-Detail": ["B6"],  # Closing surplus = opening + income + distributions
    "Summary": ["B2"],  # Reported closing surplus references the detail tab
}

#: Forbidden leftover text (lowercased substrings) — stale notes / internal words.
FORBIDDEN_TEXT: tuple[str, ...] = (
    "todo",
    "fixme",
    "tbd",
    "pending",
    "reviewer decision",
    "do not ship",
    "internal only",
    "draft - not final",
    "claude",
    "codex",
    "chatgpt",
    "copilot",
)

#: On the Trial-Balance tab, these columns hold debit / credit amounts and the
#: final row is expected to tie out (sum(debit) == sum(credit)).
TRIAL_BALANCE_SHEET = "Trial-Balance"
TB_DEBIT_COL = "B"
TB_CREDIT_COL = "C"

#: Lineage expectation: cells on the evidence tab are raw *inputs* (literals);
#: cells on the detail tab are *derived* (formulas). A reversed direction —
#: evidence holding a formula, or a detail driver holding a literal — is a FLAG.
EVIDENCE_SHEET = "Evidence"
EVIDENCE_INPUT_CELLS = ["B2", "B3", "B4"]
DETAIL_DRIVER_SHEET = "Surplus-Detail"
DETAIL_DRIVER_CELLS = ["B6"]


# --------------------------------------------------------------------------- #
# Checks
# --------------------------------------------------------------------------- #
@check("expected_formula")
def check_expected_formula(ctx: WorkbookContext) -> list[Finding]:
    """Cells that should be formula-driven must hold a formula, not a literal.

    Catches quiet manual overrides where a reviewer typed a number on top of a
    computed total.
    """
    out: list[Finding] = []
    for sheet, cells in EXPECTED_FORMULA_CELLS.items():
        if sheet not in ctx.sheetnames:
            continue
        for cell in cells:
            v = ctx.formula(sheet, cell)
            loc = f"{sheet}!{cell}"
            if _is_formula(v):
                out.append(
                    Finding("expected_formula", Status.PASS, loc, f"formula present: {v}")
                )
            else:
                out.append(
                    Finding(
                        "expected_formula",
                        Status.FAIL,
                        loc,
                        f"expected a formula but found hardcoded value: {v!r}",
                    )
                )
    return out


@check("debit_credit_balance")
def check_debit_credit_balance(ctx: WorkbookContext) -> list[Finding]:
    """Trial balance must tie out: sum(debit) == sum(credit).

    Sums the debit and credit columns from the cached values and compares them
    within a small tolerance. An out-of-balance trial balance is a FAIL.
    """
    if TRIAL_BALANCE_SHEET not in ctx.sheetnames:
        return []
    ws = ctx.value_wb[TRIAL_BALANCE_SHEET]
    debit = credit = 0.0
    # Row 1 is the header; data rows follow until the first empty label cell.
    for r in range(2, ws.max_row + 1):
        label = ws[f"A{r}"].value
        if label is None or str(label).strip().lower().startswith("total"):
            continue
        d = ws[f"{TB_DEBIT_COL}{r}"].value or 0
        c = ws[f"{TB_CREDIT_COL}{r}"].value or 0
        if isinstance(d, (int, float)):
            debit += d
        if isinstance(c, (int, float)):
            credit += c
    loc = f"{TRIAL_BALANCE_SHEET}!{TB_DEBIT_COL}:{TB_CREDIT_COL}"
    if abs(debit - credit) < 1e-6:
        return [
            Finding(
                "debit_credit_balance",
                Status.PASS,
                loc,
                f"trial balance ties out (debit==credit=={debit:g})",
            )
        ]
    return [
        Finding(
            "debit_credit_balance",
            Status.FAIL,
            loc,
            f"trial balance does NOT tie out: debit={debit:g} vs credit={credit:g} "
            f"(diff={debit - credit:g})",
        )
    ]


@check("forbidden_text")
def check_forbidden_text(ctx: WorkbookContext) -> list[Finding]:
    """Scan every cell for stale review notes / internal / process words.

    Any hit is a FLAG (human-review signal) rather than a hard failure: the
    number may be fine, but the leftover note must never reach a deliverable.
    """
    out: list[Finding] = []
    for sheet, coord, v in ctx.iter_formula_cells():
        if not isinstance(v, str):
            continue
        low = v.lower()
        for term in FORBIDDEN_TEXT:
            if term in low:
                out.append(
                    Finding(
                        "forbidden_text",
                        Status.FLAG,
                        f"{sheet}!{coord}",
                        f"forbidden text {term!r} in cell: {v!r}",
                    )
                )
                break  # one finding per cell is enough
    if not out:
        out.append(
            Finding("forbidden_text", Status.PASS, "-", "no forbidden text found")
        )
    return out


@check("lineage_direction")
def check_lineage_direction(ctx: WorkbookContext) -> list[Finding]:
    """Lineage must flow inputs -> formulas.

    Evidence cells are raw *inputs* and must be literals (a formula there means
    evidence is being *derived*, which inverts the audit trail). Detail driver
    cells are *outputs* and must be formulas. Either inversion is a FLAG.
    """
    out: list[Finding] = []

    if EVIDENCE_SHEET in ctx.sheetnames:
        for cell in EVIDENCE_INPUT_CELLS:
            v = ctx.formula(EVIDENCE_SHEET, cell)
            loc = f"{EVIDENCE_SHEET}!{cell}"
            if _is_formula(v):
                out.append(
                    Finding(
                        "lineage_direction",
                        Status.FLAG,
                        loc,
                        f"evidence input should be a literal but is a formula: {v}",
                    )
                )

    if DETAIL_DRIVER_SHEET in ctx.sheetnames:
        for cell in DETAIL_DRIVER_CELLS:
            v = ctx.formula(DETAIL_DRIVER_SHEET, cell)
            loc = f"{DETAIL_DRIVER_SHEET}!{cell}"
            if v is not None and not _is_formula(v):
                out.append(
                    Finding(
                        "lineage_direction",
                        Status.FLAG,
                        loc,
                        f"detail driver should be a formula but is a literal: {v!r}",
                    )
                )

    if not out:
        out.append(
            Finding(
                "lineage_direction",
                Status.PASS,
                "-",
                "lineage flows inputs -> formulas",
            )
        )
    return out


@check("cap_logic_leftover")
def check_cap_logic_leftover(ctx: WorkbookContext) -> list[Finding]:
    """Detect leftover MIN/MAX 'cap' logic that should have been replaced.

    A scaling / reference cell that still wraps a ``MAX(0, ...)`` or
    ``MIN(...)`` cap from an earlier draft is a FLAG: the cap was a working
    assumption, not a final presentation.
    """
    out: list[Finding] = []
    pat = re.compile(r"\b(MIN|MAX)\s*\(", re.IGNORECASE)
    for sheet, coord, v in ctx.iter_formula_cells():
        if _is_formula(v) and pat.search(v):
            out.append(
                Finding(
                    "cap_logic_leftover",
                    Status.FLAG,
                    f"{sheet}!{coord}",
                    f"leftover MIN/MAX cap logic: {v}",
                )
            )
    if not out:
        out.append(
            Finding("cap_logic_leftover", Status.PASS, "-", "no MIN/MAX cap logic found")
        )
    return out


def _closing_surplus_value(ctx: WorkbookContext) -> Optional[float]:
    """Resolve the workbook's closing surplus as a number.

    openpyxl only exposes a *cached* value for a formula cell if Excel itself
    last saved one. Workbooks produced programmatically (as the demo corpus is)
    have no cache, so we fall back to evaluating the well-known closing-surplus
    relationship ``B6 = B3 + B4 + B5`` directly from the literal detail inputs.

    Resolution order:
      1. cached value at ``Surplus-Detail!B6`` (if Excel computed one), else
      2. the literal at ``B6`` (the hardcoded-total defect case), else
      3. ``B3 + B4 + B5`` from the literal inputs.
    """
    cached = ctx.value("Surplus-Detail", "B6")
    if isinstance(cached, (int, float)):
        return float(cached)
    raw = ctx.formula("Surplus-Detail", "B6")
    if isinstance(raw, (int, float)):  # hardcoded literal in B6
        return float(raw)
    total = 0.0
    have_any = False
    for cell in ("B3", "B4", "B5"):
        part = ctx.formula("Surplus-Detail", cell)
        if isinstance(part, (int, float)):
            total += part
            have_any = True
    return total if have_any else None


@check("json_tieout")
def check_json_tieout(ctx: WorkbookContext) -> list[Finding]:
    """Cross-check a sibling JSON export against the workbook, if present.

    The JSON export is expected to carry a ``closing_surplus`` that matches the
    workbook's resolved closing surplus (see :func:`_closing_surplus_value`). A
    mismatch is a FAIL — the published JSON disagrees with the workbook it was
    derived from.
    """
    if ctx.json_data is None:
        return []
    reported = ctx.json_data.get("closing_surplus")
    if reported is None:
        return [
            Finding(
                "json_tieout",
                Status.FLAG,
                "json:closing_surplus",
                "JSON export present but has no 'closing_surplus' key",
            )
        ]
    wb_value = _closing_surplus_value(ctx)
    if wb_value is None:
        return [
            Finding(
                "json_tieout",
                Status.FLAG,
                "json:closing_surplus",
                "could not resolve workbook closing surplus to compare against JSON",
            )
        ]
    if isinstance(reported, (int, float)) and abs(float(reported) - float(wb_value)) < 1e-6:
        return [
            Finding(
                "json_tieout",
                Status.PASS,
                "json:closing_surplus",
                f"JSON closing_surplus matches workbook ({reported:g})",
            )
        ]
    return [
        Finding(
            "json_tieout",
            Status.FAIL,
            "json:closing_surplus",
            f"JSON closing_surplus={reported} != workbook closing surplus={wb_value:g}",
        )
    ]


# --------------------------------------------------------------------------- #
# Runner
# --------------------------------------------------------------------------- #
def _load_context(path: Path) -> WorkbookContext:
    """Open ``path`` read-only (twice) and attach any sibling JSON export."""
    formula_wb = load_workbook(path, data_only=False, read_only=True)
    value_wb = load_workbook(path, data_only=True, read_only=True)
    json_path = path.with_suffix(".json")
    json_data = None
    if json_path.exists():
        try:
            json_data = json.loads(json_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            json_data = None
    return WorkbookContext(
        path=path, formula_wb=formula_wb, value_wb=value_wb, json_data=json_data
    )


def validate_workbook(path: Path | str) -> WorkbookReport:
    """Run the full :data:`REGISTRY` over a single workbook (read-only).

    Parameters
    ----------
    path:
        Path to a ``.xlsx`` workbook. A sibling ``<stem>.json`` is auto-loaded
        if present.

    Returns
    -------
    WorkbookReport
        All findings plus a rolled-up :class:`Verdict`.
    """
    path = Path(path)
    ctx = _load_context(path)
    report = WorkbookReport(workbook=path.name)
    try:
        for _rule_id, fn in REGISTRY:
            report.findings.extend(fn(ctx))
    finally:
        ctx.formula_wb.close()
        ctx.value_wb.close()
    return report


def validate_folder(folder: Path | str) -> list[WorkbookReport]:
    """Validate every ``.xlsx`` in ``folder`` (non-recursive), sorted by name."""
    folder = Path(folder)
    reports: list[WorkbookReport] = []
    for path in sorted(folder.glob("*.xlsx")):
        if path.name.startswith("~$"):  # skip Excel lock files
            continue
        reports.append(validate_workbook(path))
    return reports


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #
def overall_verdict(reports: list[WorkbookReport]) -> Verdict:
    """Roll a list of workbook reports into one overall verdict."""
    if any(r.verdict is Verdict.FAIL for r in reports):
        return Verdict.FAIL
    if any(r.verdict is Verdict.REVIEW for r in reports):
        return Verdict.REVIEW
    return Verdict.PASS


def build_json_report(reports: list[WorkbookReport]) -> dict:
    """Assemble the structured JSON report for a run."""
    return {
        "tool": "validation_engine",
        "version": __import__("validation_engine").__version__,
        "generated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "overall_verdict": overall_verdict(reports).value,
        "workbook_count": len(reports),
        "registry": [rule_id for rule_id, _ in REGISTRY],
        "reports": [r.to_dict() for r in reports],
    }


_STATUS_ICON = {Status.PASS: "PASS", Status.FAIL: "FAIL", Status.FLAG: "FLAG"}
_VERDICT_ICON = {Verdict.PASS: "PASS", Verdict.REVIEW: "REVIEW", Verdict.FAIL: "FAIL"}


def build_markdown_report(reports: list[WorkbookReport]) -> str:
    """Render a human-readable markdown report for a run."""
    lines: list[str] = []
    overall = overall_verdict(reports)
    lines.append("# Validation Report")
    lines.append("")
    lines.append(f"**Overall verdict:** {_VERDICT_ICON[overall]}")
    lines.append("")
    lines.append(f"Workbooks validated: {len(reports)}")
    lines.append("")
    lines.append("| Workbook | Verdict | PASS | FAIL | FLAG |")
    lines.append("| --- | --- | --: | --: | --: |")
    for r in reports:
        c = r.counts()
        lines.append(
            f"| {r.workbook} | {_VERDICT_ICON[r.verdict]} | "
            f"{c['PASS']} | {c['FAIL']} | {c['FLAG']} |"
        )
    lines.append("")
    for r in reports:
        lines.append(f"## {r.workbook} — {_VERDICT_ICON[r.verdict]}")
        lines.append("")
        actionable = [f for f in r.findings if f.status is not Status.PASS]
        if not actionable:
            lines.append("All checks passed.")
            lines.append("")
            continue
        lines.append("| Status | Rule | Location | Message |")
        lines.append("| --- | --- | --- | --- |")
        for f in actionable:
            msg = f.message.replace("|", "\\|")
            lines.append(
                f"| {_STATUS_ICON[f.status]} | {f.rule} | {f.location} | {msg} |"
            )
        lines.append("")
    return "\n".join(lines)

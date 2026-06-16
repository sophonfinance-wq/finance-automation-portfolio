"""
Synthetic workbook generator (FICTIONAL DATA ONLY).
===================================================

Builds a corpus of fictional, surplus-style workbooks for the validation
engine to run against. One workbook is **clean** (should PASS every rule); the
others each plant **exactly one** defect so every rule in the registry has a
matching positive test case.

Determinism
-----------
All synthetic figures come from the stdlib :mod:`random` module seeded with a
fixed value (:data:`SEED`), so the generated corpus is byte-stable run to run
and the tests are reproducible.

Confidentiality
---------------
Every entity name is obviously fake ("Demo Holdings LLC", "Maple Fund LP",
"Birchwood Op Co", ...). No real people, entities, EINs, paths or figures
appear anywhere.

Each workbook has four tabs:

- ``Surplus-Detail``  — opening + income + distributions => closing (a formula).
- ``Summary``         — references the detail tab.
- ``Trial-Balance``   — debit / credit columns that must tie out.
- ``Evidence``        — raw input cells (literals) feeding the detail tab.
"""

from __future__ import annotations

import json
import random
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

#: Fixed RNG seed — guarantees a reproducible corpus.
SEED = 20240614

#: Obviously-fictional entity names.
FICTIONAL_ENTITIES: tuple[str, ...] = (
    "Demo Holdings LLC",
    "Maple Fund LP",
    "Birchwood Op Co",
    "Cedar Ridge Trust",
    "Harborview Partners LP",
    "Sandbox Capital LLC",
)


@dataclass(frozen=True)
class Defect:
    """Describes a single planted defect (one per rule)."""

    key: str  # short slug used in the filename
    rule: str  # rule id the defect should trip
    label: str  # human description


#: One defect per rule in the registry (plus a clean baseline below).
DEFECTS: tuple[Defect, ...] = (
    Defect("hardcoded_total", "expected_formula", "closing surplus hardcoded"),
    Defect("unbalanced_tb", "debit_credit_balance", "trial balance off by a plug"),
    Defect("stale_note", "forbidden_text", "leftover TODO review note"),
    Defect("reversed_lineage", "lineage_direction", "evidence cell holds a formula"),
    Defect("cap_leftover", "cap_logic_leftover", "leftover MAX() cap on summary"),
    Defect("json_mismatch", "json_tieout", "JSON export disagrees with workbook"),
)


def _amounts(rng: random.Random) -> tuple[int, int, int]:
    """Return (opening, income, distributions) — fictional, seeded figures."""
    opening = rng.randrange(800, 2000, 50)
    income = rng.randrange(100, 600, 25)
    distributions = -rng.randrange(50, 300, 25)
    return opening, income, distributions


def _build_workbook(
    entity: str, rng: random.Random, defect: Defect | None
) -> tuple[Workbook, dict]:
    """Build one workbook (and its JSON-export payload).

    Parameters
    ----------
    entity:
        Fictional entity name to stamp on the workbook.
    rng:
        Seeded RNG for figures.
    defect:
        The single defect to plant, or ``None`` for a clean workbook.

    Returns
    -------
    (Workbook, dict)
        The openpyxl workbook and the dict that will become its ``.json`` export.
    """
    key = defect.key if defect else "clean"
    opening, income, distributions = _amounts(rng)
    closing = opening + income + distributions

    wb = Workbook()

    # --- Surplus-Detail -------------------------------------------------- #
    sd = wb.active
    sd.title = "Surplus-Detail"
    sd["A1"], sd["B1"] = f"{entity} — Surplus Detail", "FY2024 (USD)"
    sd["A2"], sd["B2"] = "Entity", entity
    sd["A3"], sd["B3"] = "Opening surplus", opening
    sd["A4"], sd["B4"] = "Current-year income", income
    sd["A5"], sd["B5"] = "Distributions", distributions
    sd["A6"] = "Closing surplus"
    # Closing surplus SHOULD be a formula. The hardcoded-total defect breaks it.
    if key == "hardcoded_total":
        sd["B6"] = closing  # literal instead of formula => expected_formula FAIL
    else:
        sd["B6"] = "=B3+B4+B5"

    if key == "stale_note":
        sd["D2"] = "TODO: reviewer decision pending on FY2024 income"

    # --- Summary --------------------------------------------------------- #
    summ = wb.create_sheet("Summary")
    summ["A1"], summ["B1"] = "Reported closing surplus", "FY2024 (USD)"
    summ["A2"] = "Closing surplus"
    if key == "cap_leftover":
        # Leftover MAX() cap from an earlier draft => cap_logic_leftover FLAG.
        summ["B2"] = "=MAX(0, 'Surplus-Detail'!B6)"
    else:
        summ["B2"] = "='Surplus-Detail'!B6"

    # --- Trial-Balance --------------------------------------------------- #
    tb = wb.create_sheet("Trial-Balance")
    tb["A1"], tb["B1"], tb["C1"] = "Account", "Debit", "Credit"
    # A small, self-balancing set of fictional accounts.
    cash = rng.randrange(200, 900, 25)
    expense = rng.randrange(50, 400, 25)
    revenue = cash + expense  # credits balance debits exactly
    tb["A2"], tb["B2"], tb["C2"] = "Cash", cash, 0
    tb["A3"], tb["B3"], tb["C3"] = "Operating expense", expense, 0
    tb["A4"], tb["B4"], tb["C4"] = "Revenue", 0, revenue
    if key == "unbalanced_tb":
        # Introduce a plug so debits != credits => debit_credit_balance FAIL.
        tb["B3"] = expense + rng.randrange(10, 40, 5)

    # --- Evidence -------------------------------------------------------- #
    ev = wb.create_sheet("Evidence")
    ev["A1"], ev["B1"] = "Source figure", "Value"
    ev["A2"], ev["B2"] = "Opening (per prior workpaper)", opening
    ev["A3"], ev["B3"] = "Income (per ledger extract)", income
    ev["A4"], ev["B4"] = "Distributions (per cash log)", distributions
    if key == "reversed_lineage":
        # Evidence input should be a raw literal; a formula inverts lineage.
        ev["B2"] = "=Opening+0"  # formula on an evidence input => lineage FLAG

    # --- JSON export payload -------------------------------------------- #
    json_closing = closing
    if key == "json_mismatch":
        json_closing = closing + 25  # published JSON disagrees => json_tieout FAIL
    payload = {
        "entity": entity,
        "fiscal_year": "FY2024",
        "currency": "USD",
        "closing_surplus": json_closing,
        "source_workbook": f"{entity.replace(' ', '_')}.xlsx",
    }
    return wb, payload


def generate_corpus(out_dir: Path | str, *, seed: int = SEED) -> list[Path]:
    """Generate the full corpus into ``out_dir`` (created if needed).

    Writes one clean workbook plus one workbook per planted defect, each with a
    sibling ``.json`` export. Returns the list of written ``.xlsx`` paths.
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    # Start from a clean slate so stale artifacts from a previous run (or older
    # demo schema) cannot contaminate the corpus.
    for stale in (*out.glob("*.xlsx"), *out.glob("*.json")):
        if not stale.name.startswith("~$"):
            stale.unlink()

    rng = random.Random(seed)

    plans: list[tuple[str, Defect | None]] = [(FICTIONAL_ENTITIES[0], None)]
    for i, defect in enumerate(DEFECTS, start=1):
        entity = FICTIONAL_ENTITIES[i % len(FICTIONAL_ENTITIES)]
        plans.append((entity, defect))

    written: list[Path] = []
    for entity, defect in plans:
        key = defect.key if defect else "clean"
        wb, payload = _build_workbook(entity, rng, defect)
        stem = f"{key}__{entity.replace(' ', '_').replace('—', '-')}"
        xlsx_path = out / f"{stem}.xlsx"
        wb.save(xlsx_path)
        (out / f"{stem}.json").write_text(
            json.dumps(payload, indent=2), encoding="utf-8"
        )
        written.append(xlsx_path)
    return written


if __name__ == "__main__":
    here = Path(__file__).resolve().parent.parent
    target = here / "samples"
    paths = generate_corpus(target)
    print(f"Wrote {len(paths)} fictional workbook(s) to {target}:")
    for p in paths:
        print("  -", p.name)
    print("\nNow run:  python -m validation_engine ./samples")

"""End-to-end and behavioural tests for the Triangulate Orchestrator.

Coverage maps directly to the system's required guarantees:

* the flow runs end-to-end and emits all four artifacts;
* the Reviewer (and every read-only role) cannot mutate the workpaper
  (separation of duties enforced at the type level *and* by digest check);
* severity classification works and reconciliation ranks/de-dupes findings;
* a DEFECTIVE workpaper is caught (verdict FAIL) and a CLEAN one PASSes;
* the data is deterministic; the AnthropicReviewer stub never needs a key.
"""

from __future__ import annotations

import json
import os

import pytest

from triangulate import (
    Severity,
    TriangulateOrchestrator,
    WorkpaperMutationError,
)
from triangulate.generate import (
    build_sample,
    make_clean_workpaper,
    make_defective_workpaper,
    write_xlsx,
)
from triangulate.formula import FormulaError, evaluate
from triangulate.model import AuthoritySource, Finding
from triangulate.orchestrator import SeparationOfDutiesError
from triangulate.reconcile import (
    HumanGate,
    VerdictStatus,
    reconcile,
    severity_breakdown,
)
from triangulate.roles.preparer import DemoPreparer
from triangulate.roles.reviewer import (
    AdversarialReviewer,
    AnthropicReviewer,
    MockLLMReviewer,
)


# --------------------------------------------------------------------------- #
# End-to-end                                                                  #
# --------------------------------------------------------------------------- #
def test_pipeline_runs_end_to_end_on_defective():
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective"))
    result = orch.run()
    assert result.workpaper.entity  # built something
    assert result.builder_memo  # memo produced
    assert result.qa_summary  # QA trail produced
    # The defective sample must NOT pass.
    assert result.verdict.status is VerdictStatus.FAIL


def test_defective_is_caught_clean_passes():
    defective = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    clean = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean")).run()

    assert defective.verdict.status is VerdictStatus.FAIL
    assert not defective.verdict.passed

    assert clean.verdict.status is VerdictStatus.PASS
    assert clean.verdict.passed
    # A clean run has zero Critical/High findings.
    assert clean.verdict.severity_counts["Critical"] == 0
    assert clean.verdict.severity_counts["High"] == 0


def test_artifacts_written(tmp_path):
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective"))
    result = orch.run()
    paths = orch.write_artifacts(result, str(tmp_path))

    for title in ("Builder Memo", "Fix Packet", "Change Log", "QA Summary"):
        assert title in paths
        assert os.path.exists(paths[title])
        assert os.path.getsize(paths[title]) > 0

    # Verdict JSON is valid and reflects the FAIL outcome.
    with open(paths["Verdict (JSON)"], encoding="utf-8") as handle:
        verdict = json.load(handle)
    assert verdict["status"] == "FAIL"
    assert "findings" in verdict


# --------------------------------------------------------------------------- #
# Separation of duties                                                         #
# --------------------------------------------------------------------------- #
def test_readonly_view_blocks_attribute_mutation():
    wp = make_clean_workpaper()
    view = wp.frozen_snapshot()
    with pytest.raises(WorkpaperMutationError):
        view.engagement = "tampered"  # type: ignore[misc]


def test_readonly_view_isolates_underlying_workpaper():
    wp = make_clean_workpaper()
    view = wp.frozen_snapshot()
    original_digest = wp.digest()
    # Mutating a cell pulled from the view must not touch the real workpaper.
    cell = view.get("B2")
    assert cell is not None
    cell.value = -999_999.0
    assert wp.digest() == original_digest
    assert wp.get("B2").value != -999_999.0


def test_reviewer_cannot_mutate_workpaper():
    """The Reviewer only ever returns findings; digest is unchanged."""
    wp = make_defective_workpaper()
    before = wp.digest()
    reviewer = AdversarialReviewer(MockLLMReviewer())
    findings = reviewer.review(wp.frozen_snapshot())
    assert wp.digest() == before
    assert all(isinstance(f, Finding) for f in findings)


def test_orchestrator_detects_a_cheating_readonly_role():
    """If a read-only role mutates the live workpaper, the run aborts."""

    class CheatingReviewer(AdversarialReviewer):
        name = "Reviewer:Cheater"

        def review(self, view):  # noqa: ANN001
            # Reach past the read-only contract to the *live* workpaper the
            # orchestrator holds -- simulating a control breach.
            self._captured_wp.notes.append("illegitimate edit")
            return []

    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean"))
    cheater = CheatingReviewer()

    # Hand the cheater a reference to the live workpaper after build.
    wp = orch.preparer.build()
    cheater._captured_wp = wp  # type: ignore[attr-defined]
    orch.preparer.build = lambda: wp  # type: ignore[assignment]
    orch.reviewer = cheater

    with pytest.raises(SeparationOfDutiesError):
        orch.run()


# --------------------------------------------------------------------------- #
# Severity taxonomy + reconciliation                                          #
# --------------------------------------------------------------------------- #
def test_severity_ordering_and_lookup():
    assert Severity.CRITICAL > Severity.HIGH > Severity.MEDIUM > Severity.LOW
    assert Severity.from_name("critical") is Severity.CRITICAL
    assert Severity.from_name("Low").label == "Low"
    with pytest.raises(ValueError):
        Severity.from_name("bogus")


def test_severity_breakdown_has_all_buckets():
    counts = severity_breakdown([])
    assert counts == {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}


def test_reconcile_dedupes_and_prefers_higher_authority():
    low_auth = Finding(
        code="DUP", cell_ref="B7", severity=Severity.HIGH,
        message="low authority", raised_by="X",
        authority=AuthoritySource.AI_ASSUMPTION,
    )
    high_auth = Finding(
        code="DUP", cell_ref="B7", severity=Severity.HIGH,
        message="high authority", raised_by="Y",
        authority=AuthoritySource.SIGNED_PRIOR_YEAR,
    )
    ranked = reconcile([low_auth, high_auth])
    assert len(ranked) == 1  # de-duplicated on (code, cell_ref)
    assert ranked[0].message == "high authority"  # higher authority wins


def test_reconcile_sorts_most_severe_first():
    findings = [
        Finding("A", "B1", Severity.LOW, "low", "X"),
        Finding("B", "B2", Severity.CRITICAL, "crit", "X"),
        Finding("C", "B3", Severity.MEDIUM, "med", "X"),
    ]
    ranked = reconcile(findings)
    assert [f.severity for f in ranked] == [
        Severity.CRITICAL, Severity.MEDIUM, Severity.LOW
    ]


def test_human_gate_policy():
    gate = HumanGate()
    crit = [Finding("X", "B1", Severity.CRITICAL, "m", "r")]
    high = [Finding("X", "B1", Severity.HIGH, "m", "r")]
    low = [Finding("X", "B1", Severity.LOW, "m", "r")]
    assert gate.decide(crit).status is VerdictStatus.FAIL
    assert gate.decide(high).status is VerdictStatus.FLAG
    assert gate.decide(low).status is VerdictStatus.PASS
    assert gate.decide([]).status is VerdictStatus.PASS


# --------------------------------------------------------------------------- #
# Reviewer detects the seeded defects                                         #
# --------------------------------------------------------------------------- #
def test_reviewer_flags_tie_out_and_assumptions_on_defective():
    wp = make_defective_workpaper()
    findings = MockLLMReviewer().generate_findings(wp.frozen_snapshot())
    codes = {f.code for f in findings}
    assert "TIE_OUT_MISMATCH" in codes
    assert "UNSUPPORTED_AI_ASSUMPTION" in codes
    assert "PROCESS_LANGUAGE_LEAK" in codes
    # At least one Critical finding exists on the defective sample.
    assert any(f.severity is Severity.CRITICAL for f in findings)


def test_reviewer_clean_has_no_critical_or_high():
    wp = make_clean_workpaper()
    findings = MockLLMReviewer().generate_findings(wp.frozen_snapshot())
    assert all(f.severity < Severity.HIGH for f in findings)


# --------------------------------------------------------------------------- #
# Determinism                                                                  #
# --------------------------------------------------------------------------- #
def test_generator_is_deterministic():
    a = make_defective_workpaper(seed=42).to_dict()
    b = make_defective_workpaper(seed=42).to_dict()
    assert a == b


def test_pipeline_is_deterministic():
    r1 = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    r2 = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    assert r1.verdict.to_dict() == r2.verdict.to_dict()


# --------------------------------------------------------------------------- #
# Formula evaluator (safe, no eval)                                           #
# --------------------------------------------------------------------------- #
def test_formula_evaluator_basic():
    assert evaluate("=B2+B3", {"B2": 10, "B3": 5}) == 15
    assert evaluate("=B2*B3-B4", {"B2": 10, "B3": 5, "B4": 4}) == 46
    assert evaluate("=(B2+B3)*B4", {"B2": 1, "B3": 2, "B4": 3}) == 9


def test_formula_evaluator_rejects_bad_input():
    with pytest.raises(FormulaError):
        evaluate("=B2+", {"B2": 1})
    with pytest.raises(FormulaError):
        evaluate("=B9", {"B2": 1})  # missing cell


# --------------------------------------------------------------------------- #
# Real-LLM-ready stub stays inert (no key/network)                            #
# --------------------------------------------------------------------------- #
def test_anthropic_reviewer_is_inert_without_key(monkeypatch):
    # No key in the environment -> must raise clearly *before* any network call.
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    reviewer = AnthropicReviewer()  # constructs with no key, no network
    assert reviewer.model == "claude-opus-4-8"  # defaults to the latest Opus
    wp = make_clean_workpaper()
    with pytest.raises(RuntimeError, match="ANTHROPIC_API_KEY"):
        reviewer.generate_findings(wp.frozen_snapshot())
    # And it carries a real guardrailed system prompt.
    assert "Reviewer" in reviewer.SYSTEM_PROMPT
    assert "NEVER" in reviewer.SYSTEM_PROMPT


# --------------------------------------------------------------------------- #
# Sample factory + optional xlsx                                              #
# --------------------------------------------------------------------------- #
def test_build_sample_factory():
    assert build_sample("clean").get("B5") is not None
    assert build_sample("defective").get("B5") is not None
    with pytest.raises(ValueError):
        build_sample("nonsense")


def test_write_xlsx_roundtrip(tmp_path):
    wp = make_clean_workpaper()
    path = write_xlsx(wp, str(tmp_path / "wp.xlsx"))
    assert os.path.exists(path)
    from openpyxl import load_workbook

    book = load_workbook(path)
    sheet = book["Workpaper"]
    assert sheet.cell(row=1, column=1).value == "Ref"

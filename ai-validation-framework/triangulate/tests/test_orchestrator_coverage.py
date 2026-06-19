"""Coverage tests for the orchestrator pipeline and the synthetic generator.
"""

from __future__ import annotations

import json
import os

import pytest

from triangulate import (
    SeparationOfDutiesError,
    TriangulateOrchestrator,
    VerdictStatus,
)
from triangulate.generate import (
    build_sample,
    make_clean_workpaper,
    make_defective_workpaper,
    write_xlsx,
)
from triangulate.model import Severity
from triangulate.roles.preparer import DemoPreparer


# --------------------------------------------------------------------------- #
# generate.py                                                                  #
# --------------------------------------------------------------------------- #
def test_clean_workpaper_ties_out():
    wp = make_clean_workpaper()
    b2, b3, b4 = (wp.get(r).value for r in ("B2", "B3", "B4"))
    assert wp.get("B5").value == b2 + b3 + b4


def test_clean_has_no_ai_assumption_cells():
    wp = make_clean_workpaper()
    assert all(c.source.name != "AI_ASSUMPTION" for c in wp.ordered_cells())


def test_defective_total_does_not_tie_out():
    wp = make_defective_workpaper()
    b2, b3, b4 = (wp.get(r).value for r in ("B2", "B3", "B4"))
    # Defect 1: stated total is the true total minus 1000.
    assert wp.get("B5").value == pytest.approx(b2 + b3 + b4 - 1000.0)


def test_defective_has_ai_assumption_cells():
    wp = make_defective_workpaper()
    assert any(c.source.name == "AI_ASSUMPTION" for c in wp.ordered_cells())


def test_defective_b7_has_no_formula():
    # Defect 2: tax cell is hard-coded with no formula backing.
    assert make_defective_workpaper().get("B7").formula is None


def test_generator_deterministic_for_same_seed():
    assert make_defective_workpaper(seed=99).to_dict() == make_defective_workpaper(seed=99).to_dict()
    assert make_clean_workpaper(seed=99).to_dict() == make_clean_workpaper(seed=99).to_dict()


@pytest.mark.parametrize("seed_a, seed_b", [(1, 2), (100, 200), (2024, 2025)])
def test_generator_varies_with_seed(seed_a, seed_b):
    a = make_clean_workpaper(seed=seed_a)
    b = make_clean_workpaper(seed=seed_b)
    # Different seeds choose different driver figures -> different digests.
    assert a.digest() != b.digest()


@pytest.mark.parametrize("kind", ["clean", "defective"])
def test_build_sample_known_kinds(kind):
    wp = build_sample(kind)
    assert wp.get("B5") is not None
    assert wp.get("B7") is not None
    assert wp.get("B8") is not None


@pytest.mark.parametrize("kind", ["", "CLEAN", "Defective", "bogus", "1065"])
def test_build_sample_rejects_unknown_kind(kind):
    with pytest.raises(ValueError):
        build_sample(kind)


def test_build_sample_passes_seed_through():
    assert build_sample("clean", seed=5).to_dict() == make_clean_workpaper(5).to_dict()


def test_clean_has_eight_cells_b2_to_b8():
    wp = make_clean_workpaper()
    assert {c.ref for c in wp.ordered_cells()} == {"B2", "B3", "B4", "B5", "B6", "B7", "B8"}


# --------------------------------------------------------------------------- #
# write_xlsx round trip                                                        #
# --------------------------------------------------------------------------- #
def test_write_xlsx_creates_file_with_header(tmp_path):
    wp = make_clean_workpaper()
    path = write_xlsx(wp, str(tmp_path / "wp.xlsx"))
    assert os.path.exists(path)
    from openpyxl import load_workbook

    sheet = load_workbook(path)["Workpaper"]
    assert [sheet.cell(row=1, column=c).value for c in range(1, 6)] == [
        "Ref", "Label", "Value", "Formula", "Source"
    ]


def test_write_xlsx_writes_all_cells(tmp_path):
    wp = make_clean_workpaper()
    path = write_xlsx(wp, str(tmp_path / "wp.xlsx"))
    from openpyxl import load_workbook

    sheet = load_workbook(path)["Workpaper"]
    refs_in_sheet = {
        sheet.cell(row=r, column=1).value
        for r in range(2, 2 + len(wp.cells))
    }
    assert {c.ref for c in wp.ordered_cells()}.issubset(refs_in_sheet)


# --------------------------------------------------------------------------- #
# Orchestrator end-to-end                                                      #
# --------------------------------------------------------------------------- #
def test_pipeline_defective_fails():
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    assert result.verdict.status is VerdictStatus.FAIL
    assert not result.verdict.passed


def test_pipeline_clean_passes():
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean")).run()
    assert result.verdict.status is VerdictStatus.PASS
    assert result.verdict.passed
    assert result.verdict.severity_counts["Critical"] == 0
    assert result.verdict.severity_counts["High"] == 0


def test_pipeline_result_has_all_artifacts():
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    assert result.builder_memo
    assert result.qa_summary
    assert result.change_log
    assert result.workpaper.entity


def test_fix_packet_is_only_critical_and_high():
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    assert result.fix_packet  # defective produces actionable items
    assert all(
        f.severity in (Severity.CRITICAL, Severity.HIGH) for f in result.fix_packet
    )


def test_clean_fix_packet_is_empty():
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean")).run()
    assert result.fix_packet == []


def test_qa_summary_mentions_read_only_confirmation():
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean")).run()
    joined = "\n".join(result.qa_summary)
    assert "read-only confirmed" in joined
    assert "HumanGate verdict" in joined


def test_pipeline_without_specialist_records_skip():
    result = TriangulateOrchestrator(
        preparer=DemoPreparer(kind="clean"), use_specialist=False
    ).run()
    assert result.change_log == ["Specialist step skipped (use_specialist=False)."]


def test_pipeline_is_deterministic():
    a = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    b = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    assert a.verdict.to_dict() == b.verdict.to_dict()


def test_pipeline_does_not_mutate_workpaper_via_readonly_roles():
    # The pipeline must complete without raising a SeparationOfDutiesError,
    # i.e. every read-only role left the digest unchanged.
    result = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective")).run()
    assert result.verdict.status is VerdictStatus.FAIL


def test_orchestrator_default_roles_are_wired():
    orch = TriangulateOrchestrator()
    assert orch.preparer is not None
    assert orch.reviewer is not None
    assert orch.specialist is not None
    assert orch.auditor is not None
    assert orch.human_gate is not None


# --------------------------------------------------------------------------- #
# Artifact emission                                                            #
# --------------------------------------------------------------------------- #
def test_write_artifacts_emits_all_files(tmp_path):
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective"))
    result = orch.run()
    paths = orch.write_artifacts(result, str(tmp_path))
    for title in ("Builder Memo", "Fix Packet", "Change Log", "QA Summary", "Verdict (JSON)"):
        assert title in paths
        assert os.path.exists(paths[title])
        assert os.path.getsize(paths[title]) > 0


def test_write_artifacts_verdict_json_is_valid_and_fail(tmp_path):
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="defective"))
    result = orch.run()
    paths = orch.write_artifacts(result, str(tmp_path))
    with open(paths["Verdict (JSON)"], encoding="utf-8") as handle:
        verdict = json.load(handle)
    assert verdict["status"] == "FAIL"
    assert "findings" in verdict


def test_write_artifacts_clean_fix_packet_says_nothing_to_fix(tmp_path):
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean"))
    result = orch.run()
    paths = orch.write_artifacts(result, str(tmp_path))
    with open(paths["Fix Packet"], encoding="utf-8") as handle:
        body = handle.read()
    assert "nothing to fix" in body.lower()


def test_write_artifacts_records_paths_on_result(tmp_path):
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean"))
    result = orch.run()
    paths = orch.write_artifacts(result, str(tmp_path))
    assert result.artifact_paths == paths


def test_write_artifacts_paths_are_absolute(tmp_path):
    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean"))
    result = orch.run()
    paths = orch.write_artifacts(result, str(tmp_path))
    assert all(os.path.isabs(p) for p in paths.values())


# --------------------------------------------------------------------------- #
# Separation-of-duties enforcement                                            #
# --------------------------------------------------------------------------- #
def test_orchestrator_aborts_when_readonly_role_mutates():
    from triangulate.roles.reviewer import AdversarialReviewer

    class CheatingReviewer(AdversarialReviewer):
        name = "Reviewer:Cheater"

        def review(self, view):  # noqa: ANN001
            self._captured_wp.notes.append("illegitimate edit")
            return []

    orch = TriangulateOrchestrator(preparer=DemoPreparer(kind="clean"))
    cheater = CheatingReviewer()
    wp = orch.preparer.build()
    cheater._captured_wp = wp  # type: ignore[attr-defined]
    orch.preparer.build = lambda: wp  # type: ignore[assignment]
    orch.reviewer = cheater

    with pytest.raises(SeparationOfDutiesError):
        orch.run()

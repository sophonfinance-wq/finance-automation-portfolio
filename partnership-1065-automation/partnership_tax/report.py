"""Report writers for the partnership 1065 automation demo."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .money import fmt
from .model import FormLine, TaxPackage


def write_outputs(package: TaxPackage, out_dir: Path, write_xlsx: bool = True) -> list[Path]:
    """Write Markdown, JSON, and optional XLSX outputs."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    workpapers = out_dir / "tax_workpapers.md"
    workpapers.write_text(tax_workpapers_markdown(package), encoding="utf-8")
    written.append(workpapers)

    checks = out_dir / "review_checks.md"
    checks.write_text(review_checks_markdown(package), encoding="utf-8")
    written.append(checks)

    preview = out_dir / "form_1065_preview.json"
    preview.write_text(json.dumps(form_preview_json(package), indent=2), encoding="utf-8")
    written.append(preview)

    if write_xlsx:
        written.append(write_support_workbook(package, out_dir / "1065_supporting_package.xlsx"))
    return written


def tax_workpapers_markdown(package: TaxPackage) -> str:
    """Render the 1065 support package as Markdown."""
    src = package.source
    lines: list[str] = []
    a = lines.append
    a(f"# Partnership 1065 Tax Workpapers - {src.year} (FICTIONAL)")
    a("")
    a("> Fully synthetic demonstration data. No real taxpayer, partner, EIN, or client figure is included.")
    a("")
    a(f"**Partnership:** {src.partnership_name}")
    a(f"**EIN:** {src.ein}")
    a(f"**Package status:** {'READY FOR REVIEW' if package.ready else 'REVIEW NEEDED'}")
    a("")
    a("## 1. AI-assisted source intake")
    a("")
    a("| Source area | What the system extracts |")
    a("|---|---|")
    a("| Trial balance / P&L | income, deductions, book depreciation, ordinary income drivers |")
    a("| Balance sheet | cash, receivables, property, liabilities, capital tie-out |")
    a("| Member capital accounts | beginning capital, contributions, distributions, partner percentages |")
    a("| Syndication cost support | nondeductible book/tax adjustment |")
    a("| Return PDF / line map | 1065, Schedule K, Schedule L, M-1, M-2, and K-1 output targets |")
    a("")
    a("## 2. Book to tax bridge")
    a("")
    a("| Item | Amount | Source |")
    a("|---|---:|---|")
    a(f"| Book income per workpapers | {fmt(package.book_income_cents)} | Trial balance / P&L |")
    for adj in src.book_tax_adjustments:
        a(f"| {adj.description} | {fmt(adj.amount_cents)} | `{adj.source_id}` |")
    a(f"| **Ordinary business income** | **{fmt(package.ordinary_income_cents)}** | Schedule K line 1 |")
    a("")
    a("## 3. Form 1065 line map")
    a("")
    a("| Form | Line | Description | Amount | Source IDs |")
    a("|---|---|---|---:|---|")
    for line in package.form_lines:
        a(
            f"| {line.form} | {line.line} | {line.description} | "
            f"{fmt(line.amount_cents)} | {', '.join(f'`{s}`' for s in line.source_ids)} |"
        )
    a("")
    a("## 4. Partner K-1 allocation preview")
    a("")
    a("| Partner | Ordinary income | BOY capital | Contributions | Distributions | EOY capital |")
    a("|---|---:|---:|---:|---:|---:|")
    for alloc in package.partner_allocations:
        a(
            f"| {alloc.partner_name} | {fmt(alloc.ordinary_income_cents)} | "
            f"{fmt(alloc.beginning_capital_cents)} | {fmt(alloc.contributions_cents)} | "
            f"{fmt(alloc.distributions_cents)} | {fmt(alloc.ending_capital_cents)} |"
        )
    a("")
    a("## 5. Source index")
    a("")
    a("| Source ID | Tab / area | Cell | Label | Amount | Note |")
    a("|---|---|---|---|---:|---|")
    for rec in src.source_records:
        a(
            f"| `{rec.source_id}` | {rec.tab} | {rec.cell} | {rec.label} | "
            f"{fmt(rec.amount_cents)} | {rec.note} |"
        )
    a("")
    a("## 6. CEO / partner-ready summary")
    a("")
    a(
        "The system generated the 1065 workpaper bridge, mapped the workpapers to return lines, "
        "allocated Schedule K income to the fictional partners, and ran review checks before "
        "marking the package ready for review."
    )
    a("")
    return "\n".join(lines)


def review_checks_markdown(package: TaxPackage) -> str:
    """Render reviewer checks as Markdown."""
    lines = ["# Partnership 1065 Review Checks (FICTIONAL)", ""]
    lines.append("| Check | Description | Expected | Actual | Difference | Status | Source |")
    lines.append("|---|---|---:|---:|---:|:---:|---|")
    for check in package.checks:
        lines.append(
            f"| {check.check_id} | {check.description} | {_check_value(check, check.expected_cents)} | "
            f"{_check_value(check, check.actual_cents)} | {_check_value(check, check.difference_cents)} | "
            f"{check.status} | {check.source} |"
        )
    lines.append("")
    lines.append(f"**Overall status:** {'READY' if package.ready else 'NOT READY'}")
    lines.append("")
    return "\n".join(lines)


def _check_value(check, value: int) -> str:
    return fmt(value) if check.is_money else f"{value:,}"


def form_preview_json(package: TaxPackage) -> dict[str, Any]:
    """Return a machine-readable 1065/K-1 preview."""
    return {
        "year": package.source.year,
        "partnership_name": package.source.partnership_name,
        "ein": package.source.ein,
        "status": "READY" if package.ready else "REVIEW",
        "form_lines": [
            {
                "form": line.form,
                "line": line.line,
                "description": line.description,
                "amount_cents": line.amount_cents,
                "source_ids": list(line.source_ids),
            }
            for line in package.form_lines
        ],
        "k1_allocations": [
            {
                "partner_id": alloc.partner_id,
                "partner_name": alloc.partner_name,
                "ordinary_income_cents": alloc.ordinary_income_cents,
                "ending_capital_cents": alloc.ending_capital_cents,
            }
            for alloc in package.partner_allocations
        ],
        "checks": [
            {
                "check_id": check.check_id,
                "status": check.status,
                "difference_cents": check.difference_cents,
                "description": check.description,
            }
            for check in package.checks
        ],
    }


def write_support_workbook(package: TaxPackage, out_path: Path) -> Path:
    """Write an Excel support workbook for the fictional package."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    white = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    def header(ws, values: list[str]) -> None:
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = white
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws) -> None:
        for col in ws.columns:
            width = max((len(str(cell.value)) for cell in col if cell.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 52)

    ws = wb.active
    ws.title = "Cover"
    ws.append(["Partnership 1065 Automation Demo"])
    ws["A1"].font = Font(bold=True, size=15)
    ws.append(["Status", "READY" if package.ready else "REVIEW"])
    ws.append(["Partnership", package.source.partnership_name])
    ws.append(["Year", package.source.year])
    ws.append(["Confidentiality", "Fictional data only"])
    autosize(ws)

    ws = wb.create_sheet("Form Line Map")
    header(ws, ["Form", "Line", "Description", "Amount", "Source IDs"])
    for line in package.form_lines:
        ws.append([line.form, line.line, line.description, line.amount_cents / 100, ", ".join(line.source_ids)])
    for cell in ws["D"][1:]:
        cell.number_format = "$#,##0.00;[Red]($#,##0.00);-"
    autosize(ws)

    ws = wb.create_sheet("K-1 Allocations")
    header(ws, ["Partner", "Ordinary Income", "BOY Capital", "Contributions", "Distributions", "EOY Capital"])
    for alloc in package.partner_allocations:
        ws.append([
            alloc.partner_name,
            alloc.ordinary_income_cents / 100,
            alloc.beginning_capital_cents / 100,
            alloc.contributions_cents / 100,
            alloc.distributions_cents / 100,
            alloc.ending_capital_cents / 100,
        ])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = "$#,##0.00;[Red]($#,##0.00);-"
    autosize(ws)

    ws = wb.create_sheet("Review Checks")
    header(ws, ["Check", "Description", "Expected", "Actual", "Difference", "Status", "Source"])
    for check in package.checks:
        ws.append([
            check.check_id,
            check.description,
            check.expected_cents / 100,
            check.actual_cents / 100,
            check.difference_cents / 100,
            check.status,
            check.source,
        ])
        ws[ws.max_row][5].fill = ok_fill if check.status == "OK" else fail_fill
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "$#,##0.00;[Red]($#,##0.00);-"
    autosize(ws)

    ws = wb.create_sheet("Source Index")
    header(ws, ["Source ID", "Tab / Area", "Cell", "Label", "Amount", "Note"])
    for rec in package.source.source_records:
        ws.append([rec.source_id, rec.tab, rec.cell, rec.label, rec.amount_cents / 100, rec.note])
    for cell in ws["E"]:
        cell.number_format = "$#,##0.00;[Red]($#,##0.00);-"
    autosize(ws)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        if sheet.max_row > 1:
            for cell in sheet[1]:
                cell.font = bold if cell.fill != header_fill else white

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def line_amount(package: TaxPackage, form: str, line: str) -> int:
    """Return a line amount from the package."""
    for form_line in package.form_lines:
        if form_line.form == form and form_line.line == line:
            return form_line.amount_cents
    raise KeyError(f"{form} line {line}")
